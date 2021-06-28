import logging

import datetime
import openpyxl
from openpyxl.styles import Font

from pyexcel_xlsx import get_data as xlsx_get
from django.db import transaction
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.db.models import Q

from brand.models import Brand
from categories.models import Category
from products.models import Product, ParentProduct, ParentProductTaxMapping, ProductHSN, ParentProductCategory, Tax, \
    Repackaging, DestinationRepackagingCostMapping, ProductTaxMapping

logger = logging.getLogger(__name__)
info_logger = logging.getLogger('file-info')
error_logger = logging.getLogger('file-error')


class SetMasterData(object):
    """
    It will call the functions mentioned in UploadMasterData class as per the condition of header_list
    """

    @classmethod
    def set_master_data(cls, excel_file_data_list):
        UploadMasterData.set_inactive_status(excel_file_data_list)
        UploadMasterData.set_parent_data(excel_file_data_list)
        UploadMasterData.set_child_parent(excel_file_data_list)
        UploadMasterData.set_sub_brand_and_brand(excel_file_data_list)
        UploadMasterData.set_sub_category_and_category(excel_file_data_list)
        UploadMasterData.set_child_data(excel_file_data_list)


class UploadMasterData(object):
    """
    This function will be used for following operations:
    a)Set the Status to "Deactivated" for a Product
    b)Mapping of "Sub Brand" to "Brand"
    c)Mapping of "Sub Category" to "Category"
    d)Set the data for "Parent SKU"
    e)Mapping of Child SKU to Parent SKU
    f)Set the Child SKU Data
    """

    @classmethod
    def set_inactive_status(cls, excel_file_data_list):
        try:
            count = 0
            info_logger.info("Method Start to set Inactive status from excel file")
            for row in excel_file_data_list:
                if row['status'] == 'deactivated':
                    count += 1
                    product = Product.objects.filter(product_sku=row['sku_id'])
                    if 'mrp' in row.keys() and not row['mrp'] == '':
                        product.update(status='deactivated', product_mrp=row['mrp'])
                    else:
                        product.update(status='deactivated')
                else:
                    continue
            info_logger.info("Set Inactive Status function called -> Inactive row id count :" + str(count))
            info_logger.info("Method Complete to set the Inactive status from excel file")
        except Exception as e:
            error_logger.info(f"Something went wrong, while working with 'Set Inactive Status Functionality' + {str(e)}")

    @classmethod
    def set_sub_brand_and_brand(cls, excel_file_data_list):
        """
            Updating Brand & Sub Brand
        """
        try:
            count = 0
            row_num = 1
            sub_brand = []
            info_logger.info('Method Start to set the Sub-brand to Brand mapping from excel file')
            for row in excel_file_data_list:
                count += 1
                row_num += 1
                try:
                    if 'sub_brand_id' in row.keys() and not row['sub_brand_id'] == '':
                        if row['sub_brand_id'] == row['brand_id']:
                            continue
                        else:
                            Brand.objects.filter(id=row['sub_brand_id']).update(brand_parent=row['brand_id'])
                except:
                    sub_brand.append(str(row_num))
            info_logger.info("Total row executed :" + str(count))
            info_logger.info("Sub brand is not updated in these row :" + str(sub_brand))
            info_logger.info("Method complete to set the Sub-Brand to Brand mapping from csv file")
        except Exception as e:
            error_logger.info(
                f"Something went wrong, while working with 'Set Sub Brand and Brand Functionality' + {str(e)}")

    @classmethod
    def set_sub_category_and_category(cls, excel_file_data_list):
        try:
            count = 0
            row_num = 1
            sub_category = []
            info_logger.info("Method Start to set the Sub-Category to Category mapping from excel file")
            for row in excel_file_data_list:
                count += 1
                row_num += 1
                try:
                    if 'sub_category_id' in row.keys() and not row['sub_category_id'] == '':
                        if row['sub_category_id'] == row['category_id']:
                            continue
                        else:
                            Category.objects.filter(id=row['sub_category_id']).update(category_parent=row['category_id'])
                except:
                    sub_category.append(str(row_num))
            info_logger.info("Total row executed :" + str(count))
            info_logger.info("Sub Category is not updated in these row :" + str(sub_category))
            info_logger.info("Method Complete to set the Sub-Category to Category mapping from excel file")
        except Exception as e:
            error_logger.info(
                f"Something went wrong, while working with 'Set Sub Category and Category Functionality' + {str(e)}")

    @classmethod
    def set_parent_data(cls, excel_file_data_list):
        try:
            count = 0
            row_num = 1
            parent_data = []
            info_logger.info("Method Start to set the data for Parent SKU")
            for row in excel_file_data_list:
                row_num += 1
                if not row['status'] == 'deactivated':
                    count += 1
                    try:
                        parent_product = ParentProduct.objects.filter(parent_id=row['parent_id'])

                        fields = ['product_type', 'hsn', 'tax_1(gst)', 'tax_2(cess)',
                                  'tax_3(surcharge)', 'brand_case_size', 'inner_case_size', 'brand_id',
                                  'sub_brand_id', 'category_id', 'sub_category_id', 'is_ptr_applicable', 'ptr_type',
                                  'ptr_percent', 'is_ars_applicable', 'max_inventory_in_days', 'is_lead_time_applicable']

                        available_fields = []
                        for col in fields:
                            if col in row.keys():
                                if row[col] != '':
                                    available_fields.append(col)

                        for col in available_fields:

                            if col == 'product_type':
                                parent_product.update(product_type=row['product_type'])

                            if col == 'hsn':
                                parent_product.update(product_hsn=ProductHSN.objects.filter(product_hsn_code=row['hsn']).last())

                            if col == 'tax_1(gst)':
                                tax = Tax.objects.filter(tax_name=row['tax_1(gst)'])
                                ParentProductTaxMapping.objects.filter(parent_product=parent_product[0].id,
                                                                       tax__tax_type='gst').update(tax=tax[0])
                                if 'sku_id' in row.keys() and row['sku_id'] != '':
                                    product = Product.objects.filter(product_sku=row['sku_id'])
                                    ProductTaxMapping.objects.filter(product=product[0].id, tax__tax_type='gst').\
                                        update(tax=tax[0])

                            if col == 'tax_2(cess)':
                                tax = Tax.objects.filter(tax_name=row['tax_2(cess)'])
                                ParentProductTaxMapping.objects.filter(parent_product=parent_product[0].id,
                                                                       tax__tax_type='cess').update(tax=tax[0])
                                if 'sku_id' in row.keys() and row['sku_id'] != '':
                                    product = Product.objects.filter(product_sku=row['sku_id'])
                                    ProductTaxMapping.objects.filter(product=product[0].id, tax__tax_type='cess').\
                                        update(tax=tax[0])

                            if col == 'tax_3(surcharge)':
                                tax = Tax.objects.filter(tax_name=row['tax_3(surcharge)'])
                                ParentProductTaxMapping.objects.filter(parent_product=parent_product[0].id, tax__tax_type='surcharge').update(
                                    tax=tax[0])
                                if 'sku_id' in row.keys() and row['sku_id'] != '':
                                    product = Product.objects.filter(product_sku=row['sku_id'])
                                    ProductTaxMapping.objects.filter(product=product[0].id, tax__tax_type='surcharge').\
                                        update(tax=tax[0])

                            if col == 'inner_case_size':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update\
                                    (inner_case_size=row['inner_case_size'])

                            if col == 'sub_category_id':
                                if row['sub_category_id'] == row['category_id']:
                                    continue
                                else:
                                    ParentProductCategory.objects.filter(parent_product=parent_product[0].id).update(
                                        category=Category.objects.filter(id=row['sub_category_id']).last())

                            if col == 'sub_brand_id':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update(
                                    parent_brand=Brand.objects.filter(id=row['sub_brand_id']).last())

                            if col == 'is_ptr_applicable':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update\
                                    (is_ptr_applicable=True if row['is_ptr_applicable'].lower() == 'yes' else False)

                            if col == 'ptr_type':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update \
                                    (ptr_type=None if not row['is_ptr_applicable'].lower() == 'yes' else ParentProduct.PTR_TYPE_CHOICES.MARK_UP
                                              if row['ptr_type'].lower() == 'mark up'else ParentProduct.PTR_TYPE_CHOICES.MARK_DOWN)

                            if col == 'ptr_percent':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update \
                                    (ptr_percent=None if not row['is_ptr_applicable'].lower() == 'yes' else row['ptr_percent'])

                            if col == 'is_ars_applicable':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update \
                                    (is_ars_applicable=True if row['is_ars_applicable'].lower() == 'yes' else False)

                            if col == 'max_inventory_in_days':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update \
                                    (max_inventory=row['max_inventory_in_days'])

                            if col == 'is_lead_time_applicable':
                                ParentProduct.objects.filter(parent_id=row['parent_id']).update \
                                    (is_lead_time_applicable=True if row['is_lead_time_applicable'].lower() == 'yes' else False)

                    except Exception as e:
                        parent_data.append(str(row_num) + ' ' + str(e))
            info_logger.info("Total row executed :" + str(count))
            info_logger.info("Some Error Found in these rows, while working with Parent Data Functionality :" + str(parent_data))
            info_logger.info("Method Complete to set the data for Parent SKU")
        except Exception as e:
            error_logger.info(
                f"Something went wrong, while working with 'Set Parent Data Functionality' + {str(e)}")

    @classmethod
    def set_child_parent(cls, excel_file_data_list):
        try:
            info_logger.info("Method Start to set the Child to Parent mapping from excel file")
            count = 0
            row_num = 1
            set_child = []
            for row in excel_file_data_list:
                row_num += 1
                if not row['status'] == 'deactivated':
                    count += 1
                    try:
                        child_pro = Product.objects.filter(product_sku=row['sku_id'])
                        parent_pro = ParentProduct.objects.filter(parent_id=row['parent_id']).last()
                        child_pro.update(parent_product=parent_pro)
                    except:
                        set_child.append(str(row_num))

            info_logger.info("Total row executed :" + str(count))
            info_logger.info("Child SKU is not exist in these row :" + str(set_child))
            info_logger.info("Method complete to set the Child to Parent mapping from excel file")
        except Exception as e:
            error_logger.info(
                f"Something went wrong, while working with 'Set Child Parent Functionality' + {str(e)}")

    @classmethod
    def set_child_data(cls, excel_file_data_list):
        try:
            info_logger.info("Method Start to set the Child data from excel file")
            count = 0
            row_num = 1
            child_data = []
            for row in excel_file_data_list:
                row_num += 1
                count += 1
                try:
                    child_product = Product.objects.filter(product_sku=row['sku_id'])
                    child_product.update(product_name=row['sku_name'], status=row['status'])
                    fields = ['ean', 'mrp', 'weight_unit', 'weight_value']
                    available_fields = []
                    for col in fields:
                        if col in row.keys() and row[col] != '':
                            available_fields.append(col)

                    for col in available_fields:
                        if col == 'ean':
                            child_product.update(product_ean_code=row['ean'])
                        if col == 'mrp':
                            child_product.update(product_mrp=row['mrp'])
                        if col == 'weight_value':
                            child_product.update(weight_value=row['weight_value'])

                    if 'repackaging_type' in row.keys() and row['repackaging_type'] == 'destination':
                        DestinationRepackagingCostMapping.objects.filter(destination=child_product[0].id)\
                                                                 .update(raw_material=row['raw_material'],
                                                                         wastage=row['wastage'],
                                                                         fumigation=row['fumigation'],
                                                                         label_printing=row['label_printing'],
                                                                         packing_labour=row['packing_labour'],
                                                                         primary_pm_cost=row['primary_pm_cost'],
                                                                         secondary_pm_cost=row['secondary_pm_cost'],
                                                                         final_fg_cost=row['final_fg_cost'],
                                                                         conversion_cost=row['conversion_cost'])
                except Exception as e:
                    child_data.append(str(row_num) + ' ' + str(e))
            info_logger.info("Total row executed :" + str(count))
            info_logger.info("Some error found in these row while working with Child data Functionality:" + str(child_data))
            info_logger.info("Script Complete to set the Child data")
        except Exception as e:
            error_logger.info(
                f"Something went wrong, while working with 'Set Child Data Functionality' + {str(e)}")


class DownloadMasterData(object):
    """
        This function will be used for following operations:
        a)return an Sample Excel File in xlsx format which can be used for uploading the master_data
        b)return an Sample Excel File in xlsx format which can be used for Status to "Deactivated" for a Product
        c)return an Sample Excel File in xlsx format which can be used for Mapping of "Sub Brand" to "Brand"
        d)return an Sample Excel File in xlsx format which can be used for Mapping of "Sub Category" to "Category"
        e)return an Sample Excel File in xlsx format which can be used for Set the data for "Parent SKU"
        f)return an Sample Excel File in xlsx format which can be used for Mapping of Child SKU to Parent SKU
        g)return an Sample Excel File in xlsx format which can be used for Set the Child SKU Data
        """

    def response_workbook(xlx_filename):

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response[
            'Content-Disposition'] = 'attachment; filename={date}-{xlx_filename}.xlsx'.format(
            date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'), xlx_filename=xlx_filename,
        )
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Users'

        return response, workbook, worksheet

    @classmethod
    def set_master_data_sample_excel_file(cls):
        pass

    @classmethod
    def set_inactive_status_sample_excel_file(cls):
        pass

    @classmethod
    def brand_sub_brand_mapping_sample_excel_file(cls):

        response, workbook, worksheet = DownloadMasterData.response_workbook("subBrand-BrandMappingSample")
        # Sheet header, first row
        row_num = 1

        columns = ['brand_id', 'brand_name', 'sub_brand_id', 'sub_brand_name', ]
        mandatory_columns = ['brand_id', 'brand_name']

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            if column_title in mandatory_columns:
                ft = Font(color="000000FF", name="Arial", b=True)
            else:
                ft = Font(color="00000000", name="Arial", b=True)
            cell.font = ft

        brands = Brand.objects.values('id', 'brand_name', 'brand_parent_id', 'brand_parent__brand_name')
        for brand in brands:
            row = []
            if brand['brand_parent_id']:
                row.append(brand['brand_parent_id'])
                row.append(brand['brand_parent__brand_name'])
                row.append(brand['id'])
                row.append(brand['brand_name'])

            else:
                row.append(brand['id'])
                row.append(brand['brand_name'])
                row.append(brand['brand_parent_id'])
                row.append(brand['brand_parent__brand_name'])

            row_num += 1
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        info_logger.info("Brand and Sub Brand Mapping Sample Excel File has been Successfully Downloaded")
        return response

    @classmethod
    def category_sub_category_mapping_sample_excel_file(cls):
        response, workbook, worksheet = DownloadMasterData.response_workbook("subCategory-CategorySample")

        # Sheet header, first row
        row_num = 1

        columns = ['category_id', 'category_name', 'sub_category_id', 'sub_category_name', ]
        mandatory_columns = ['category_id', 'category_name']

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            if column_title in mandatory_columns:
                ft = Font(color="000000FF", name="Arial", b=True)
            else:
                ft = Font(color="00000000", name="Arial", b=True)
            cell.font = ft

        categories = Category.objects.values('id', 'category_name', 'category_parent_id',
                                             'category_parent__category_name')
        for category in categories:
            row = []
            if category['category_parent_id']:
                row.append(category['category_parent_id'])
                row.append(category['category_parent__category_name'])
                row.append(category['id'])
                row.append(category['category_name'])
            else:
                row.append(category['id'])
                row.append(category['category_name'])
                row.append(category['category_parent_id'])
                row.append(category['category_parent__category_name'])
            row_num += 1
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        info_logger.info("Category and Sub Category Mapping Sample Excel File has been Successfully Downloaded")
        return response

    @classmethod
    def set_child_with_parent_sample_excel_file(cls):
        pass

    @classmethod
    def set_child_data_sample_excel_file(cls):
        pass

    @classmethod
    def set_parent_data_sample_excel_file(cls):
        pass
