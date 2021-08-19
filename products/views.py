import csv
import codecs
import datetime
import math
import os
import logging
import json
import re

import boto3
from decimal import Decimal
from botocore.exceptions import ClientError
from decouple import config

import openpyxl
from openpyxl.styles import Font
from pyexcel_xlsx import get_data as xlsx_get

from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.db import transaction
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import permission_required
from admin_auto_filters.views import AutocompleteJsonView
from retailer_to_sp.models import BulkOrder, Order
from shops.models import Shop, ShopType
from addresses.models import City, Address, Pincode
from categories.models import Category
from brand.models import Brand, Vendor

from wms.models import InventoryType, WarehouseInventory, InventoryState, BinInventory, Bin, \
    WarehouseInternalInventoryChange, Out, In
from wms.common_functions import get_stock, StockMovementCSV, create_batch_id, InCommonFunctions, \
    CommonBinInventoryFunctions, CommonWarehouseInventoryFunctions, InternalInventoryChange, \
    InternalStockCorrectionChange, inventory_in_and_out_weight, get_manufacturing_date

from .forms import (
    GFProductPriceForm, ProductPriceForm, ProductsFilterForm,
    ProductsPriceFilterForm, ProductsCSVUploadForm, ProductImageForm,
    ProductCategoryMappingForm, NewProductPriceUpload, UploadParentProductAdminForm,
    UploadChildProductAdminForm, ParentProductImageForm, BulkProductVendorMapping,
    UploadMasterDataAdminForm, UploadSlabProductPriceForm, UploadPackingSkuInventoryAdminForm,
    UploadDiscountedProductPriceForm
)
from .master_data import UploadMasterData, SetMasterData
from products.models import (
    Product, ProductCategory, ProductOption,
    ProductTaxMapping, ProductVendorMapping,
    ProductImage, ProductHSN, ProductPrice,
    ParentProduct, ParentProductCategory,
    ProductSourceMapping,
    ParentProductTaxMapping, Tax, ParentProductImage,
    DestinationRepackagingCostMapping, BulkUploadForProductAttributes, Repackaging, SlabProductPrice, PriceSlab,
    ProductPackingMapping, DiscountedProductPrice
)
from products.utils import hsn_queryset
from global_config.models import GlobalConfig

from global_config.views import get_config
from pos.models import RetailerProduct

logger = logging.getLogger(__name__)

info_logger = logging.getLogger('file-info')
error_logger = logging.getLogger('file-error')

from dal import autocomplete
from django.db.models import Q
from .utils import products_price_excel
from retailer_backend.utils import getStrToDate


def load_cities(request):
    """Return list of cities for specific state id

    :param request: state_id
    :return: list of cities
    """
    state_id = request.GET.get('state')
    if state_id:
        cities = City.objects.filter(state=state_id).order_by('city_name')
    else:
        cities = City.objects.none()

    return render(
        request, 'admin/products/city_dropdown_list_options.html',
        {'cities': cities}
    )


def load_sp_sr(request):
    """Return list of sp/sr for specific state and city

    :param request: state_id, city_id, sp/sr
    :return: list of sp/sr for specific state and city
    """
    state_id = request.GET.get('state_id')
    city_id = request.GET.get('city_id')
    sp_sr = request.GET.get('sp_sr')
    if sp_sr and city_id and state_id:
        shops_id = Address.objects.filter(
            city=city_id
        ).values_list('shop_name', flat=True)

        shops = Shop.objects.filter(
            pk__in=shops_id,
            shop_type=sp_sr
        ).order_by('shop_name')

        return render(
            request,
            'admin/products/shop_dropdown_list_options.html',
            {'shops': shops}
        )
    else:
        shops = Shop.objects.none()
        return render(
            request,
            'admin/products/shop_dropdown_list_options.html',
            {'shops': shops}
        )


def load_gf(request):
    """Returns list of GramFactories for specific state and city

    :param request: state_id, city_id
    :return: list of GramFactories for specific state and city
    """
    state_id = request.GET.get('state_id')
    city_id = request.GET.get('city_id')
    if city_id and state_id:
        shops_id = Address.objects.filter(
            city=city_id
        ).values_list('shop_name', flat=True)

        shoptype = ShopType.objects.filter(shop_type="gf")

        shops = Shop.objects.filter(
            pk__in=shops_id, shop_type__in=shoptype
        ).order_by('shop_name')

        return render(
            request,
            'admin/products/shop_dropdown_list_options.html',
            {'shops': shops}
        )
    else:
        shops = Shop.objects.none()
        return render(
            request,
            'admin/products/shop_dropdown_list_options.html',
            {'shops': shops}
        )


def load_brands(request):
    """Returns brands for specific category_id

    :param request: category_id
    :return: list of brands
    """
    id = request.GET.get('category_id')
    if id:
        from urllib.parse import unquote

        id = list(filter(None, [x.strip() for x in unquote(id).split(',')]))

        category_id = Category.objects.filter(id__in=id).values('id')

        product_id = ProductCategory.objects.filter(
            category__in=category_id
        ).values_list('product')

        product_brand = Product.objects.filter(
            id__in=product_id
        ).values_list('product_brand')

        brands = Brand.objects.filter(
            id__in=product_brand
        ).order_by('brand_name')

        return render(
            request,
            'admin/products/brand_dropdown_list_options.html',
            {'brands': brands}
        )
    else:
        brands = Brand.objects.none()
        return render(
            request,
            'admin/products/brand_dropdown_list_options.html',
            {'brands': brands}
        )


class SpSrProductPrice(View):

    def validate_row(self, first_row, row):
        if not re.match("^\d{0,8}(\.\d{1,4})?$", row[4]) or not row[4]:
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[4], first_row[4]))
        if not re.match("^\d{0,8}(\.\d{1,4})?$", row[5]) or not row[5]:
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[5], first_row[5]))
        if not re.match("^\d{0,8}(\.\d{1,4})?$", row[6]) or not row[6]:
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[6], first_row[6]))
        if not re.match("^\d{0,8}(\.\d{1,4})?$", row[7]) or not row[7]:
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[7], first_row[7]))
        if row[8] and not re.match("^\d{0,8}(\.\d{1,4})?$", row[8]):
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[8], first_row[8]))
        if row[9] and not re.match("^\d{0,8}(\.\d{1,4})?$", row[9]):
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[9], first_row[9]))
        if (row[0] and not re.match("^[\d]*$", row[0])) or not row[0]:
            raise Exception("{} - Please enter a valid {}"
                            "".format(row[0], first_row[0]))

    def create_product_price(self, request, file, shops, city, start_date,
                             end_date, sp_sr):
        try:
            with transaction.atomic():
                reader = csv.reader(codecs.iterdecode(file, 'utf-8', errors='ignore'))
                first_row = next(reader)
                for row_id, row in enumerate(reader):

                    if (row[4] and row[5] and row[6] and row[7]):
                        self.validate_row(first_row, row)
                        for shop in shops:
                            ProductPrice.objects.create(
                                product_id=row[0], city_id=city,
                                mrp=float(row[4]), shop_id=shop.id,
                                price_to_retailer=float(row[7]),
                                price_to_service_partner=float(row[5]),
                                price_to_super_retailer=float(row[6]),
                                cash_discount=float(row[8]) if row[8] else 0,
                                loyalty_incentive=float(row[9]) if row[9] else 0,
                                start_date=start_date, end_date=end_date,
                                approval_status=ProductPrice.APPROVAL_PENDING)

                    elif (row[4] or row[5] or row[6] or row[7]):
                        raise Exception("Please enter all the prices")
                    else:
                        continue

                messages.success(request, 'Price uploaded successfully')

        except Exception as e:
            messages.error(request, "{} at Row[{}] for {}"
                                    "".format(e, row_id + 2, row[1]))

    def get(self, request):
        form = ProductPriceForm(initial={'sp_sr_list': Shop.objects.none()})
        return render(request, 'admin/products/productpriceupload.html',
                      {'form': form})

    def post(self, request):
        form = ProductPriceForm(request.POST, request.FILES)

        if form.is_valid():
            file = form.cleaned_data.get('file')
            city = form.cleaned_data.get('city').id
            start_date = form.cleaned_data.get('start_date_time')
            end_date = form.cleaned_data.get('end_date_time')
            sp_sr = form.cleaned_data.get('sp_sr_choice').shop_type
            shops = form.cleaned_data.get('sp_sr_list')

            self.create_product_price(request, file, shops, city,
                                      start_date, end_date, sp_sr)

        return render(
            request,
            'admin/products/productpriceupload.html',
            {'form': form}
        )


def validate_row(row):
    if not re.match("^\d{0,8}(\.\d{1,4})?$", row[4]) or not row[4]:
        raise Exception("{} - Please enter a valid {}"
                        "".format(row[4], "mrp"))
    if not re.match("^\d{0,8}(\.\d{1,4})?$", row[5]) or not row[5]:
        raise Exception("{} - Please enter a valid {}"
                        "".format(row[5], "price_to_service_partner"))
    if not re.match("^\d{0,8}(\.\d{1,4})?$", row[6]) or not row[6]:
        raise Exception("{} - Please enter a valid {}"
                        "".format(row[6], "price_to_super_retailer"))
    if not re.match("^\d{0,8}(\.\d{1,4})?$", row[7]) or not row[7]:
        raise Exception("{} - Please enter a valid {}"
                        "".format(row[7], "price_to_retailer"))


def gf_product_price(request):
    """CSV to product prices for GramFactory

    :param request: Form
    :return: product prices for GramFactory
    """
    if request.method == 'POST':
        form = GFProductPriceForm(request.POST, request.FILES)

        if form.errors:
            return render(
                request,
                'admin/products/gfproductpriceupload.html',
                {'form': form}
            )

        if form.is_valid():
            file = form.cleaned_data.get('file')
            city = form.cleaned_data.get('city').id
            start_date = form.cleaned_data.get('start_date_time')
            end_date = form.cleaned_data.get('end_date_time')
            shops = form.cleaned_data.get('gf_list')
            reader = csv.reader(codecs.iterdecode(file, 'utf-8', errors='ignore'))
            first_row = next(reader)
            try:
                for row in reader:
                    validate_row(row)
                    for shop in shops:
                        product_price = ProductPrice.objects.create(
                            product_id=row[0],
                            city_id=city,
                            mrp=row[4],
                            shop_id=shop.id,
                            price_to_retailer=row[7],
                            price_to_service_partner=row[5],
                            price_to_super_retailer=row[6],
                            cash_discount=row[8],
                            loyalty_incentive=row[9],
                            start_date=start_date,
                            end_date=end_date
                        )

                messages.success(request, 'Price uploaded successfully')

            except Exception as e:
                messages.error(str(e))
                # messages.error("Something went wrong!")
            return redirect('admin:gf_productprice')

    else:
        form = GFProductPriceForm(
            initial={'gf_list': Shop.objects.none()}
        )
    return render(
        request,
        'admin/products/gfproductpriceupload.html',
        {'form': form}
    )


def products_filter_view(request):
    """Returns CSV includes products of specific category and brand

    :param request: Form
    :return: Products CSV for selected Category and Brands
    """
    if request.method == 'POST':
        form = ProductsFilterForm(request.POST, request.FILES)

        if form.errors:
            return render(
                request,
                'admin/products/productsfilter.html',
                {'filter_form': form}
            )
        if form.is_valid():
            dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
            filename = str(dt) + "product_list.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
            writer = csv.writer(response)
            writer.writerow([
                'id', 'product_name',
                'gf_code', 'product_hsn',
                'mrp', 'ptsp', 'ptsr', 'ptr', 'cash_discount', 'loyalty_incentive'
            ])
            brands = form.cleaned_data.get('brand')
            products = Product.objects.select_related(
                'product_hsn'
            ).filter(
                product_brand__in=brands
            )
            for product in products:
                writer.writerow([
                    product.id, product.product_name,
                    product.product_gf_code, product.product_hsn,
                    '', '', '', ''
                ])
            return response
    else:
        form = ProductsFilterForm()
    return render(
        request,
        'admin/products/productsfilter.html',
        {'filter_form': form}
    )


def products_price_filter_view(request):
    """Returns CSV includes product price of selected SP/SR/GF

    :param request: Form
    :return: Products CSV for selected Category and Brands
    """
    if request.method == 'POST':
        form = ProductsPriceFilterForm(request.POST, request.FILES)

        if form.errors:
            return render(
                request,
                'admin/products/productpricefilter.html',
                {'form': form}
            )
        if form.is_valid():
            city = form.cleaned_data.get('city').id
            sp_sr = form.cleaned_data.get('sp_sr_choice').shop_type
            shops = form.cleaned_data.get('sp_sr_list')
            dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
            filename = str(dt) + "product_price_list.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
            writer = csv.writer(response)

            if sp_sr == "sp":
                writer.writerow([
                    'product_id', 'product_name', 'gf_code', 'product_hsn',
                    'mrp', 'ptsp', 'ptr', 'cash_discount', 'loyalty_incentive', 'price_start_date',
                    'price_end_date', 'sp_name'
                ])
                for shop in shops:
                    products = ProductPrice.objects.select_related(
                        'product'
                    ).filter(
                        shop=shop,
                        status=True
                    ).order_by('product__product_name')
                    for product in products:
                        writer.writerow([
                            product.product_id, product.product.product_name,
                            product.product.product_gf_code,
                            product.product.product_hsn,
                            product.mrp, product.price_to_service_partner,
                            product.price_to_retailer, product.cash_discount,
                            product.loyalty_incentive, product.start_date,
                            product.end_date, product.shop.shop_name
                        ])

            if sp_sr == "sr":
                writer.writerow([
                    'product_id', 'product_name', 'gf_code', 'product_hsn',
                    'mrp', 'ptr', 'cash_discount', 'loyalty_incentive', 'price_start_date', 'price_end_date',
                    'sr_name'
                ])
                for shop in shops:
                    products = ProductPrice.objects.select_related(
                        'product'
                    ).filter(
                        shop=shop,
                        status=True
                    ).order_by('product__product_name')
                    for product in products:
                        writer.writerow([
                            product.product_id,
                            product.product.product_name,
                            product.product.product_gf_code,
                            product.product.product_hsn,
                            product.mrp, product.price_to_retailer,
                            product.cash_discount, product.loyalty_incentive,
                            product.start_date, product.end_date,
                            product.shop.shop_name
                        ])
            if sp_sr == "gf":
                writer.writerow([
                    'product_id', 'product_name', 'gf_code', 'product_hsn',
                    'mrp', 'ptsp', 'ptsr', 'ptr', 'cash_discount', 'loyalty_incentive', 'price_start_date',
                    'price_end_date', 'sr_name'
                ])
                for shop in shops:
                    products = ProductPrice.objects.select_related(
                        'product'
                    ).filter(
                        shop=shop,
                        status=True
                    ).order_by('product__product_name')
                    for product in products:
                        writer.writerow([
                            product.product_id,
                            product.product.product_name,
                            product.product.product_gf_code,
                            product.product.product_hsn,
                            product.mrp, product.price_to_service_partner,
                            product.price_to_super_retailer,
                            product.price_to_retailer, product.cash_discount,
                            product.loyalty_incentive, product.start_date,
                            product.end_date, product.shop.shop_name
                        ])
            return response

    else:
        form = ProductsPriceFilterForm()
    return render(
        request,
        'admin/products/productpricefilter.html',
        {'form': form}
    )


def products_csv_upload_view(request):
    """CSV to Product Price

    :param request: form
    :return: creates product prices
    """
    if request.method == 'POST':
        form = ProductsCSVUploadForm(request.POST, request.FILES)

        if form.errors:
            return render(
                request,
                'admin/products/productscsvupload.html',
                {'form': form}
            )

        if form.is_valid():
            file = form.cleaned_data['file']
            reader = csv.reader(codecs.iterdecode(file, 'utf-8', errors='ignore'))
            first_row = next(reader)
            for row in reader:
                try:
                    product_hsn_dt, _ = ProductHSN.objects.get_or_create(
                        product_hsn_code=row[16]
                    )
                except Exception as e:
                    logger.exception("Unable to create product HSN")
                    messages.error(request, "Unable to create product HSN")
                    return render(
                        request,
                        'admin/products/productscsvupload.html',
                        {'form': form}
                    )
                try:
                    brand = Brand.objects.get(pk=row[5])
                except Exception as e:
                    logger.exception("Brand Does not exist")
                    messages.error(request, "Brand doesn't exist for  {}".format(row[1]))
                    return render(request, 'admin/products/productscsvupload.html', {'form': form})

                try:
                    product = Product.objects.get(product_gf_code=row[3])
                except:
                    product = Product.objects.create(product_gf_code=row[3], product_brand_id=row[5])
                else:
                    product.product_brand = brand
                finally:
                    product.product_name = row[0]
                    product.product_short_description = row[1]
                    product.product_long_description = row[2]
                    product.product_ean_code = row[4]
                    product.product_inner_case_size = row[14]
                    product.product_hsn = product_hsn_dt
                    product.product_case_size = row[15]
                    product.weight_value = row[12]
                try:
                    product.save()
                except Exception as e:
                    logger.exception("Unable to save product")
                    messages.error(
                        request,
                        "unable to save product details for {}".format(row[1]))
                    return render(
                        request,
                        'admin/products/productscsvupload.html',
                        {'form': form}
                    )

                for c in row[6].split(','):
                    if c is not '':
                        try:
                            product_category, _ = ProductCategory.objects. \
                                get_or_create(
                                product=product,
                                category_id=c.strip()
                            )
                        except Exception as e:
                            logger.exception(
                                "unable to get or create product "
                                "category for category {}".format(c)
                            )
                            messages.error(
                                request, "unable to create product category "
                                         "for {}, {}".format(row[1], c)
                            )
                            return render(
                                request,
                                'admin/products/productscsvupload.html',
                                {'form': form}
                            )
                # try:
                #     productoptions, _ = ProductOption.objects.get_or_create(
                #         product=product
                #     )
                #     productoptions.size_id = row[8] if row[8] else None
                #     productoptions.color_id = row[9] if row[9] else None
                #     productoptions.fragrance_id = row[10] if row[10] else None
                #     productoptions.flavor_id = row[11] if row[11] else None
                #     productoptions.weight_id = row[12] if row[12] else None
                #     productoptions.package_size_id = row[13] if row[13] else None
                #     productoptions.save()
                # except Exception as e:
                #     logger.exception("Unable to create Product Options")
                #     messages.error(
                #         request,
                #         "Unable to create Product options "
                #         "for {}".format(row[1])
                #     )
                #     return render(
                #         request,
                #         'admin/products/productscsvupload.html',
                #         {'form': form}
                #     )
                for t in row[7].split(','):
                    if t is not '':
                        try:
                            product_tax, _ = ProductTaxMapping.objects \
                                .get_or_create(
                                product=product,
                                tax_id=t.strip()
                            )
                        except Exception as e:
                            logger.error(e)
                            messages.error(
                                request,
                                "Unable to create product tax"
                                " for {}--{}".format(row[1], t)
                            )
                            return render(
                                request,
                                'admin/products/productscsvupload.html',
                                {'form': form}
                            )
            messages.success(request, 'Products uploaded successfully')
            return redirect('admin:productscsvupload')
    else:
        form = ProductsCSVUploadForm()
    return render(
        request,
        'admin/products/productscsvupload.html',
        {'form': form}
    )


class MultiPhotoUploadView(View):
    """
    Bulk images upload with Child SKU ID as photo name
    """

    def get(self, request):
        photos_list = ProductImage.objects.all()
        return render(
            self.request,
            'admin/products/multiphotoupload.html',
            {'photos': photos_list}
        )

    def post(self, request):
        form = ProductImageForm(self.request.POST, self.request.FILES)
        if form.is_valid():
            file_name = (
                os.path.splitext(form.cleaned_data['image'].name)[0])
            product_sku = file_name.split("_")[0]
            try:
                product = Product.objects.get(product_sku=product_sku)
            except:
                data = {
                    'is_valid': False,
                    'error': True,
                    'name': 'No Product found with SKU ID <b>{}</b>'.format(product_sku),
                    'url': '#'
                }

            else:
                form_instance = form.save(commit=False)
                form_instance.product = product
                form_instance.image_name = file_name
                form_instance.save()
                data = {
                    'is_valid': True,
                    'name': form_instance.image.name,
                    'url': form_instance.image.url,
                    'product_sku': product.product_sku,
                    'product_name': product.product_name
                }
        else:
            data = {'is_valid': False}
        return JsonResponse(data)


class ParentProductMultiPhotoUploadView(View):
    """
    Bulk images upload with Parent ID as photo name
    """

    def get(self, request):
        photos_list = ParentProductImage.objects.all()
        return render(
            self.request,
            'admin/products/parentproductmultiphotoupload.html',
            {'photos': photos_list}
        )

    def post(self, request):
        form = ParentProductImageForm(self.request.POST, self.request.FILES)
        if form.is_valid():
            file_name = (
                os.path.splitext(form.cleaned_data['image'].name)[0])
            parent_id = file_name.split("_")[0]
            try:
                parent = ParentProduct.objects.get(parent_id=parent_id)
            except:
                data = {
                    'is_valid': False,
                    'error': True,
                    'name': 'No Parent Product found with Parent ID <b>{}</b>'.format(parent_id),
                    'url': '#'
                }

            else:
                form_instance = form.save(commit=False)
                form_instance.parent_product = parent
                form_instance.image_name = file_name
                form_instance.save()
                data = {
                    'is_valid': True,
                    'name': form_instance.image.name,
                    'url': form_instance.image.url,
                    'product_sku': parent.parent_id,
                    'product_name': parent.name
                }
        else:
            data = {'is_valid': False}
        return JsonResponse(data)


def export(request):
    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + "product_list.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(['id', 'product_name', 'mrp', 'ptsp', 'ptsr', 'ptr'])
    products = Product.objects.values_list('id', 'product_name')
    for product in products:
        writer.writerow([product[0], product[1], '', '', '', ''])
    return response


def products_vendor_mapping(request, pk=None):
    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + "vendor_product_list.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    try:
        writer.writerow(['parent_id', 'parent_name', 'id', 'product_name', 'sku', 'case_size', 'number_of_cases', 'mrp',
                         'brand_to_gram_price_unit', 'brand_to_gram_price'])
        vendor_products = ProductVendorMapping.objects.filter(vendor_id=int(pk), case_size__gt=0, status=True)

        for p in vendor_products:
            if p.brand_to_gram_price_unit == "Per Piece":
                writer.writerow(
                    [p.product.parent_product.parent_id, p.product.parent_name, p.product_id, p.product.product_name,
                     p.product.product_sku, p.case_size, '', p.product.product_mrp, p.brand_to_gram_price_unit,
                     p.product_price])
            else:
                writer.writerow(
                    [p.product.parent_product.parent_id, p.product.parent_name, p.product_id, p.product.product_name,
                     p.product.product_sku, p.case_size, '', p.product.product_mrp, p.brand_to_gram_price_unit,
                     p.product_price_pack])
    except:
        writer.writerow(["Make sure you have selected vendor before downloading CSV file"])
    return response


def cart_products_mapping(request, pk=None):
    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    current_time = datetime.datetime.now()
    filename = str(dt) + "cart_product_list.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    try:
        writer.writerow(['SKU', 'product_name', 'qty', 'discounted_price'])
        cart_products = ProductPrice.objects.values('product__product_sku', 'product__product_name').filter(
            seller_shop_id=int(pk), approval_status=2)
        writer.writerows(
            [(product.get('product__product_sku'), product.get('product__product_name'), '', '') for product in
             cart_products])
    except:
        writer.writerow(["Make sure you have selected seller shop before downloading CSV file"])
    return response


def cart_product_list_status(request, order_status_info):
    p = re.compile('(?<!\\\\)\'')
    order_status_info = p.sub('\"', order_status_info)
    order_status_info = json.loads(order_status_info)
    info_logger.info(f"[products/views.py]-cart_product_list_status function called for Downloading the CSV file of "
                     f"Bulk/Discounted Order Status")
    unavailable_skus = []
    for ele in order_status_info.keys():
        unavailable_skus.append(ele)
    bulk_order_obj = BulkOrder.objects.filter(cart_id=int(order_status_info['cart_id']))
    if bulk_order_obj:
        csv_file_name = bulk_order_obj.values()[0]['cart_products_csv']
    else:
        info_logger.info(f"[products/views.py:cart_product_list_status] - [cart_id : {order_status_info['cart_id']}]")

    try:
        s3 = boto3.resource('s3', aws_access_key_id=config('AWS_ACCESS_KEY_ID'),
                            aws_secret_access_key=config('AWS_SECRET_ACCESS_KEY'))
        info_logger.info(f"[products/views.py:cart_product_list_status] - Successfully connected with s3")
    except ClientError as err:
        error_logger.error(f"[products/views.py:cart_product_list_status] Failed to connect with s3 - {err}")
        raise err

    bucket = s3.Bucket(config('AWS_STORAGE_BUCKET_NAME'))
    obj = bucket.Object(key=f'media/{csv_file_name}')
    try:
        res = obj.get()
        info_logger.info(f"[products/views.py:cart_product_list_status] Successfully get the response from s3")
    except ClientError as err:
        error_logger.error(f"[products/views.py:cart_product_list_status] Failed to get the response from s3 - {err}")
        raise err

    lines = res['Body'].read()
    lines_list = lines.decode('utf-8')
    csv_data = lines_list.split('\n')
    reader = csv.reader(csv_data)
    header = csv_data[0].split(',')
    headers = []
    for ele in header:
        headers.append(ele.replace('"', ''))

    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + " - Cart_Product_List_Status.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    index = 0
    for row in reader:
        if len(row) > 0:
            if row == headers:
                writer.writerow(row + ["order_status"])
            else:
                if row[0] in unavailable_skus:
                    writer.writerow(row + [order_status_info[row[0]]])
                    index = index + 1
                else:
                    writer.writerow(row + ["Success"])
                    index = index + 1
    info_logger.info(f"[products/views.py: cart_product_list_status] - CSV for cart_product_list_status has been "
                     f"successfully downloaded with response [{response}]")
    return response


def ProductsUploadSample(request):
    filename = "products_upload_sample.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(
        ['product_name', 'product_short_description', 'product_long_description', 'product_gf_code', 'product_ean_code',
         'p_brand_id', 'p_cat_id', 'p_tax_id', 'p_size_id', 'p_color_id', 'p_fragrance_id', 'p_flavor_id',
         'weight_value(gm)', 'p_package_size_id', 'p_inner_case_size', 'p_case_size', 'product_hsn_code'])
    writer.writerow(['fortune sunflowers oil', 'Fortune Sun Lite Refined Sunflower Oil is a healthy',
                     'Fortune Sun Lite Refined Sunflower Oil is a healthy, light and nutritious oil that is simple to digest. Rich in natural vitamins, it consists mostly of poly-unsaturated fatty acids (PUFA) and is low in soaked fats. It is strong and makes you feel light and active level after heavy food.',
                     '12BBPRG00000121', '1234567890123', '1', '1', '1', '1', '1', '1', '1', '1', '1', '4', '2',
                     'HSN Code'])
    return response


def NameIDCSV(request):
    filename = "name_id.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(
        ['BRAND NAME', 'BRAND ID', 'CATEGORY NAME', 'CATEGORY ID', 'TAX NAME', 'TAX ID', 'SIZE NAME', 'SIZE ID',
         'COLOR NAME', 'COLOR ID', 'FRAGRANCE NAME', 'FRAGRANCE ID', 'FLAVOR NAME', 'FLAVOR ID', 'WEIGHT NAME',
         'WEIGHT ID', 'PACKSIZE NAME', 'PACKSIZE ID'])
    return response


class ProductPriceAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self, *args, **kwargs):
        qs = Product.objects.filter(product_type=Product.PRODUCT_TYPE_CHOICE.NORMAL)
        if self.q:
            qs = Product.objects.filter(
                Q(product_name__icontains=self.q) |
                Q(product_sku__icontains=self.q)
            )
        return qs


class ProductCategoryAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self, *args, **kwargs):
        qs = None
        if self.q:
            qs = Category.objects.filter(category_name__icontains=self.q),
            # qs = Product.objects.filter(product_name__icontains=self.q)
        return qs


def download_all_products(request):
    """Returns CSV includes products

    :param request: Form
    :return: Products CSV
    """
    products_list = Product.objects.values(
        'id', 'product_name', 'product_gf_code', 'product_hsn').all()

    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + "all_products_list.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow([
        'id', 'product_name',
        'gf_code', 'product_hsn',
        'mrp', 'ptsp', 'ptsr', 'ptr', 'cash_discount', 'loyalty_incentive'
    ])
    writer.writerows([[i['id'], i['product_name'], i['product_gf_code'],
                       i['product_hsn'], '', '', '', '']
                      for i in products_list])
    return response


def ParentProductsDownloadSampleCSV(request):
    filename = "parent_products_sample.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(
        ["Name", "Brand", "Category", "HSN", "GST", "CESS", "Surcharge", "Brand Case Size", "Inner Case Size",
         "Product Type"])
    writer.writerow(
        ["testparent2", "Nestle", "Health Care, Beverages, Grocery & Staples", "123456", "18", "12", "100", "10", "10",
         "b2c"])
    return response


def parent_product_upload(request):
    if request.method == 'POST':
        form = UploadParentProductAdminForm(request.POST, request.FILES)

        if form.errors:
            return render(request, 'admin/products/parent-product-upload.html', {'form': form})

        if form.is_valid():
            upload_file = form.cleaned_data.get('file')
            reader = csv.reader(codecs.iterdecode(upload_file, 'utf-8', errors='ignore'))
            first_row = next(reader)

            def gst_mapper(gst):
                if '0' in gst:
                    return 0
                elif '5' in gst:
                    return 5
                elif '12' in gst:
                    return 12
                elif '18' in gst:
                    return 18
                elif '28' in gst:
                    return 28

            def cess_mapper(cess):
                if '0' in cess:
                    return 0
                elif '12' in cess:
                    return 12

            try:
                for row in reader:
                    if len(row) == 0:
                        continue
                    if '' in row:
                        if (row[0] == '' and row[1] == '' and row[2] == '' and row[3] == '' and row[4] == '' and
                                row[5] == '' and row[6] == '' and row[7] == '' and row[8] == '' and row[9] == ''):
                            continue
                    parent_product = ParentProduct.objects.create(
                        name=row[0].strip(),
                        parent_brand=Brand.objects.filter(brand_name=row[1].strip()).last(),
                        product_hsn=ProductHSN.objects.filter(product_hsn_code=row[3].replace("'", '')).last(),
                        inner_case_size=int(row[7]),
                        product_type=row[8]
                    )
                    parent_product.save()
                    parent_gst = gst_mapper(row[4])
                    ParentProductTaxMapping.objects.create(
                        parent_product=parent_product,
                        tax=Tax.objects.filter(tax_type='gst', tax_percentage=parent_gst).last()
                    ).save()
                    parent_cess = cess_mapper(row[5]) if row[5] else 0
                    ParentProductTaxMapping.objects.create(
                        parent_product=parent_product,
                        tax=Tax.objects.filter(tax_type='cess', tax_percentage=parent_cess).last()
                    ).save()
                    parent_surcharge = float(row[6]) if row[6] else 0
                    if Tax.objects.filter(
                            tax_type='surcharge',
                            tax_percentage=parent_surcharge
                    ).exists():
                        ParentProductTaxMapping.objects.create(
                            parent_product=parent_product,
                            tax=Tax.objects.filter(tax_type='surcharge', tax_percentage=parent_surcharge).last()
                        ).save()
                    else:
                        new_surcharge_tax = Tax.objects.create(
                            tax_name='Surcharge - {}'.format(parent_surcharge),
                            tax_type='surcharge',
                            tax_percentage=parent_surcharge,
                            tax_start_at=datetime.datetime.now()
                        )
                        new_surcharge_tax.save()
                        ParentProductTaxMapping.objects.create(
                            parent_product=parent_product,
                            tax=new_surcharge_tax
                        ).save()
                    if Category.objects.filter(category_name=row[2].strip()).exists():
                        parent_product_category = ParentProductCategory.objects.create(
                            parent_product=parent_product,
                            category=Category.objects.filter(category_name=row[2].strip()).last()
                        )
                        parent_product_category.save()
                    else:
                        categories = row[2].split(',')
                        for cat in categories:
                            cat = cat.strip().replace("'", '')
                            parent_product_category = ParentProductCategory.objects.create(
                                parent_product=parent_product,
                                category=Category.objects.filter(category_name=cat).last()
                            )
                            parent_product_category.save()
            except Exception as e:
                return render(request, 'admin/products/parent-product-upload.html', {
                    'form': form,
                    'error': e,
                })
            return render(request, 'admin/products/parent-product-upload.html', {
                'form': form,
                'success': 'Parent Product CSV uploaded successfully !',
            })
    else:
        form = UploadParentProductAdminForm()
    return render(request, 'admin/products/parent-product-upload.html', {'form': form})


class ParentProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = ParentProduct.objects.all()

        if self.q:
            qs = qs.filter(Q(name__icontains=self.q) | Q(parent_id__icontains=self.q))

        return qs


class ProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Product.objects.all()

        if self.q:
            qs = qs.filter(name__istartswith=self.q)

        return qs


def ChildProductsDownloadSampleCSV(request):
    filename = "child_products_sample.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(["Parent Product ID", "Reason for Child SKU", "Product Name", "Product EAN Code",
                     "Product MRP", "Weight Value", "Weight Unit", "Repackaging Type", "Map Source SKU",
                     'Raw Material Cost', 'Wastage Cost', 'Fumigation Cost', 'Label Printing Cost',
                     'Packing Labour Cost', 'Primary PM Cost', 'Secondary PM Cost', "Packing SKU",
                     "Packing Sku Weight (gm) Per Unit (Qty) Destination Sku"])
    writer.writerow(["PHEAMGI0001", "Default", "TestChild1", "abcdefgh", "50", "20", "Gram", "none"])
    writer.writerow(["PHEAMGI0001", "Default", "TestChild2", "abcdefgh", "50", "20", "Gram", "source"])
    writer.writerow(["PHEAMGI0001", "Default", "TestChild3", "abcdefgh", "50", "20", "Gram", "destination",
                     "SNGSNGGMF00000016, SNGSNGGMF00000016", "10.22", "2.33", "7", "4.33", "5.33", "10.22", "5.22",
                     "BPOBLKREG00000001", "10.00"])
    return response


def product_csv_upload(request):
    if request.method == 'POST':
        form = UploadChildProductAdminForm(request.POST, request.FILES)

        if form.errors:
            return render(request, 'admin/products/child-product-upload.html', {'form': form})

        if form.is_valid():
            error = 'Something Went Wrong!'
            msg = 'Child Product CSV uploaded successfully !'
            upload_file = form.cleaned_data.get('file')
            reader = csv.reader(codecs.iterdecode(upload_file, 'utf-8', errors='ignore'))
            first_row = next(reader)

            def reason_for_child_sku_mapper(reason):
                reason = reason.lower()
                if 'default' in reason:
                    return 'default'
                elif 'mrp' in reason:
                    return 'different_mrp'
                elif 'weight' in reason:
                    return 'different_weight'
                elif 'ean' in reason:
                    return 'different_ean'
                elif 'offer' in reason:
                    return 'offer'

            try:
                with transaction.atomic():
                    for row_id, row in enumerate(reader):
                        if len(row) == 0:
                            continue
                        if '' in row:
                            if (row[0] == '' and row[1] == '' and row[2] == '' and row[3] == '' and row[4] == '' and
                                    row[5] == '' and row[6] == ''):
                                continue
                        source_map = []
                        if row[7] == 'destination':
                            for pro in row[8].split(','):
                                pro = pro.strip()
                                if pro is not '' and pro not in source_map and \
                                        Product.objects.filter(product_sku=pro, repackaging_type='source').exists():
                                    source_map.append(pro)

                        product = Product.objects.create(
                            parent_product=ParentProduct.objects.filter(parent_id=row[0]).last(),
                            reason_for_child_sku=reason_for_child_sku_mapper(row[1]),
                            product_name=row[2],
                            product_ean_code=row[3].replace("'", ''),
                            product_mrp=float(row[4]),
                            weight_value=float(row[5]),
                            weight_unit='gm' if 'gram' in row[6].lower() else 'gm',
                            repackaging_type=row[7]
                        )
                        product.save()
                        if row[7] == 'destination':
                            for sku in source_map:
                                psm = ProductSourceMapping.objects.create(
                                    destination_sku=product,
                                    source_sku=Product.objects.filter(product_sku=sku,
                                                                      repackaging_type='source').last(),
                                    status=True
                                )
                                psm.save()
                            dcm = DestinationRepackagingCostMapping.objects.create(
                                destination=product,
                                raw_material=float(row[9]),
                                wastage=float(row[10]),
                                fumigation=float(row[11]),
                                label_printing=float(row[12]),
                                packing_labour=float(row[13]),
                                primary_pm_cost=float(row[14]),
                                secondary_pm_cost=float(row[15])
                            )
                            dcm.save()
                            ProductPackingMapping.objects.create(
                                sku=product,
                                packing_sku=Product.objects.get(product_sku=row[16]),
                                packing_sku_weight_per_unit_sku=row[17]
                            )
                error = ''
            except Exception as e:
                error_logger.error(e)
                msg = ''
            return render(request, 'admin/products/child-product-upload.html', {
                'form': form,
                'success': msg,
                'error': error
            })
    else:
        form = UploadChildProductAdminForm()
    return render(request, 'admin/products/child-product-upload.html', {'form': form})


def UploadMasterDataSampleExcelFile(request, *args):
    """
    This function will return an Sample Excel File in xlsx format which can be used for uploading the master_data
    """
    category_id = request.GET['category_id']
    category = Category.objects.values('category_name').filter(id=int(category_id))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-master_data_sample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['sku_id', 'sku_name', 'parent_id', 'parent_name', 'ean', 'mrp', 'hsn',
               'weight_unit', 'weight_value', 'tax_1(gst)', 'tax_2(cess)', 'tax_3(surcharge)',
               'inner_case_size', 'brand_id', 'brand_name', 'sub_brand_id',
               'sub_brand_name', 'category_id', 'category_name', 'sub_category_id', 'sub_category_name',
               'status', 'repackaging_type', 'source_sku_id', 'source_sku_name', 'raw_material',
               'wastage', 'fumigation', 'label_printing', 'packing_labour', 'primary_pm_cost',
               'secondary_pm_cost', 'final_fg_cost', 'conversion_cost']
    mandatory_columns = ['sku_id', 'sku_name', 'parent_id', 'parent_name', 'status']

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        if column_title in mandatory_columns:
            ft = Font(color="000000FF", name="Arial", b=True)
        else:
            ft = Font(color="00000000", name="Arial", b=True)
        cell.font = ft

    products = Product.objects.values('id', 'product_sku', 'product_name', 'parent_product__parent_id',
                                      'parent_product__name', 'product_ean_code', 'product_mrp',
                                      'parent_product__product_hsn__product_hsn_code', 'weight_unit', 'weight_value',
                                      'parent_product__inner_case_size', 'parent_product__parent_brand__id',
                                      'parent_product__parent_brand__brand_name',
                                      'parent_product__parent_brand__brand_parent_id',
                                      'parent_product__parent_brand__brand_parent__brand_name',
                                      'status', 'repackaging_type') \
        .filter(
        Q(parent_product__parent_product_pro_category__category__category_name__icontains=category[0]['category_name']))

    cat = Category.objects.values('id', 'category_name', 'category_parent_id', 'category_parent__category_name').filter(
        id=int(category_id))

    for product in products:
        row = []
        tax_list = ['', '', '']
        row.append(product['product_sku'])
        row.append(product['product_name'])
        row.append(product['parent_product__parent_id'])
        row.append(product['parent_product__name'])
        row.append(product['product_ean_code'])
        row.append(product['product_mrp'])
        row.append(product['parent_product__product_hsn__product_hsn_code'])
        row.append(product['weight_unit'])
        row.append(product['weight_value'])
        taxes = ProductTaxMapping.objects.select_related('tax').filter(product=product['id'])
        for tax in taxes:
            if tax.tax.tax_type == 'gst':
                tax_list[0] = tax.tax.tax_name
            if tax.tax.tax_type == 'cess':
                tax_list[1] = tax.tax.tax_name
            if tax.tax.tax_type == 'surcharge':
                tax_list[2] = tax.tax.tax_name
        row.extend(tax_list)
        row.append(product['parent_product__inner_case_size'])

        if product['parent_product__parent_brand__brand_parent_id']:
            row.append(product['parent_product__parent_brand__brand_parent_id'])
            row.append(product['parent_product__parent_brand__brand_parent__brand_name'])
            row.append(product['parent_product__parent_brand__id'])
            row.append(product['parent_product__parent_brand__brand_name'])
        else:
            row.append(product['parent_product__parent_brand__id'])
            row.append(product['parent_product__parent_brand__brand_name'])
            row.append(product['parent_product__parent_brand__brand_parent_id'])
            row.append(product['parent_product__parent_brand__brand_parent__brand_name'])

        if cat[0]['category_parent_id']:
            row.append(cat[0]['category_parent_id'])
            row.append(cat[0]['category_parent__category_name'])
            row.append(cat[0]['id'])
            row.append(cat[0]['category_name'])
        else:
            row.append(cat[0]['id'])
            row.append(cat[0]['category_name'])
            row.append(cat[0]['category_parent_id'])
            row.append(cat[0]['category_parent__category_name'])

        row.append(product['status'])
        row.append(product['repackaging_type'])
        source_sku_name = Repackaging.objects.select_related('source_sku').filter(destination_sku=product['id'])
        source_sku_ids = []
        source_sku_names = []
        for sourceSKU in source_sku_name:
            if sourceSKU.source_sku.product_sku not in source_sku_ids:
                source_sku_ids.append(sourceSKU.source_sku.product_sku)
            if sourceSKU.source_sku.product_name not in source_sku_names:
                source_sku_names.append(sourceSKU.source_sku.product_name)
        if source_sku_ids:
            row.append(str(source_sku_ids))
        else:
            row.append('')
        if source_sku_names:
            row.append(str(source_sku_names))
        else:
            row.append('')
        costs = DestinationRepackagingCostMapping.objects.values('raw_material', 'wastage', 'fumigation',
                                                                 'label_printing', 'packing_labour', 'primary_pm_cost',
                                                                 'secondary_pm_cost', 'final_fg_cost',
                                                                 'conversion_cost').filter(destination=product['id'])
        for cost in costs:
            row.append(cost['raw_material'])
            row.append(cost['wastage'])
            row.append(cost['fumigation'])
            row.append(cost['label_printing'])
            row.append(cost['packing_labour'])
            row.append(cost['primary_pm_cost'])
            row.append(cost['secondary_pm_cost'])
            row.append(cost['final_fg_cost'])
            row.append(cost['conversion_cost'])
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    info_logger.info("Master Data Sample Excel File has been Successfully Downloaded")
    return response


def set_inactive_status_sample_excel_file(request):
    """
    This function will return an Sample Excel File in xlsx format which can be used for set Active Status
    """
    category_id = request.GET['category_id']
    category = Category.objects.values('category_name').filter(id=int(category_id))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-active_inactive_status_sample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['sku_id', 'sku_name', 'mrp', 'status', ]
    mandatory_columns = ['sku_id', 'sku_name', 'status']

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        if column_title in mandatory_columns:
            ft = Font(color="000000FF", name="Arial", b=True)
        else:
            ft = Font(color="00000000", name="Arial", b=True)
        cell.font = ft

    products = Product.objects.values('product_sku', 'product_name', 'status', 'product_mrp') \
        .filter(
        Q(parent_product__parent_product_pro_category__category__category_name__icontains=category[0]['category_name']))
    for product in products:
        row = []
        row.append(product['product_sku'])
        row.append(product['product_name'])
        row.append(product['product_mrp'])
        row.append(product['status'])
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    info_logger.info("Set Inactive Status Sample Excel File has been Successfully Downloaded")
    return response


def set_child_with_parent_sample_excel_file(request):
    """
    This function will return an Sample Excel File in xlsx format which can be used for Child and Parent Mapping
    """
    category_id = request.GET['category_id']
    category = Category.objects.values('category_name').filter(id=int(category_id))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-child_parent_mapping_data_sample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['sku_id', 'sku_name', 'parent_id', 'parent_name', 'status', ]

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        ft = Font(color="000000FF", name="Arial", b=True)
        cell.font = ft

    products = Product.objects.values('product_sku', 'product_name',
                                      'parent_product__parent_id', 'parent_product__name', 'status') \
        .filter(
        Q(parent_product__parent_product_pro_category__category__category_name__icontains=category[0]['category_name']))

    for product in products:
        row = []
        row.append(product['product_sku'])
        row.append(product['product_name'])
        row.append(product['parent_product__parent_id'])
        row.append(product['parent_product__name'])
        row.append(product['status'])
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    info_logger.info("Child Parent Mapping Sample Excel File has been Successfully Downloaded")
    return response


def get_ptr_type_text(ptr_type=None):
    if ptr_type is not None and ptr_type in ParentProduct.PTR_TYPE_CHOICES:
        return ParentProduct.PTR_TYPE_CHOICES[ptr_type]
    return ''


def set_parent_data_sample_excel_file(request, *args):
    """
    This function will return an Sample Excel File in xlsx format which can be used for set parent product data
    """
    category_id = request.GET['category_id']
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-parent_data_sample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['parent_id', 'parent_name', 'product_type', 'hsn', 'tax_1(gst)', 'tax_2(cess)', 'tax_3(surcharge)',
               'inner_case_size', 'brand_id', 'brand_name', 'sub_brand_id', 'sub_brand_name',
               'category_id', 'category_name', 'sub_category_id', 'sub_category_name', 'status',
               'is_ptr_applicable', 'ptr_type', 'ptr_percent', 'is_ars_applicable', 'max_inventory_in_days',
               'is_lead_time_applicable']

    mandatory_columns = ['parent_id', 'parent_name', 'status']

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        if column_title in mandatory_columns:
            ft = Font(color="000000FF", name="Arial", b=True)
        else:
            ft = Font(color="00000000", name="Arial", b=True)
        cell.font = ft

    parent_products = ParentProductCategory.objects.values('parent_product__id', 'parent_product__parent_id',
                                                           'parent_product__name',
                                                           'parent_product__product_type',
                                                           'parent_product__product_hsn__product_hsn_code',
                                                           'parent_product__inner_case_size',
                                                           'parent_product__parent_brand__id',
                                                           'parent_product__parent_brand__brand_name',
                                                           'parent_product__parent_brand__brand_parent_id',
                                                           'parent_product__parent_brand__brand_parent__brand_name',
                                                           'category__id', 'category__category_name',
                                                           'category__category_parent_id',
                                                           'category__category_parent__category_name',
                                                           'parent_product__status',
                                                           'parent_product__is_ptr_applicable',
                                                           'parent_product__ptr_type',
                                                           'parent_product__ptr_percent',
                                                           'parent_product__is_ars_applicable',
                                                           'parent_product__max_inventory',
                                                           'parent_product__is_lead_time_applicable',
                                                           'parent_product__discounted_life_percent').filter(
        category=int(category_id))
    for product in parent_products:
        row = []
        tax_list = ['', '', '']
        row.append(product['parent_product__parent_id'])
        row.append(product['parent_product__name'])
        row.append(product['parent_product__product_type'])
        row.append(product['parent_product__product_hsn__product_hsn_code'])
        taxes = ParentProductTaxMapping.objects.select_related('tax').filter(
            parent_product=product['parent_product__id'])
        for tax in taxes:
            if tax.tax.tax_type == 'gst':
                tax_list[0] = tax.tax.tax_name
            if tax.tax.tax_type == 'cess':
                tax_list[1] = tax.tax.tax_name
            if tax.tax.tax_type == 'surcharge':
                tax_list[2] = tax.tax.tax_name
        row.extend(tax_list)
        row.append(product['parent_product__inner_case_size'])

        if product['parent_product__parent_brand__brand_parent_id']:
            row.append(product['parent_product__parent_brand__brand_parent_id'])
            row.append(product['parent_product__parent_brand__brand_parent__brand_name'])
            row.append(product['parent_product__parent_brand__id'])
            row.append(product['parent_product__parent_brand__brand_name'])
        else:
            row.append(product['parent_product__parent_brand__id'])
            row.append(product['parent_product__parent_brand__brand_name'])
            row.append(product['parent_product__parent_brand__brand_parent_id'])
            row.append(product['parent_product__parent_brand__brand_parent__brand_name'])

        if product['category__category_parent_id']:
            row.append(product['category__category_parent_id'])
            row.append(product['category__category_parent__category_name'])
            row.append(product['category__id'])
            row.append(product['category__category_name'])
        else:
            row.append(product['category__id'])
            row.append(product['category__category_name'])
            row.append(product['category__category_parent_id'])
            row.append(product['category__category_parent__category_name'])

        if type(product['parent_product__status']):
            row.append("active")
        else:
            row.append("deactivated")
        row.append('Yes' if product['parent_product__is_ptr_applicable'] else 'No')
        row.append(get_ptr_type_text(product['parent_product__ptr_type']))
        row.append(product['parent_product__ptr_percent'])
        row.append('Yes' if product['parent_product__is_ars_applicable'] else 'No')
        row.append(product['parent_product__max_inventory'])
        row.append('Yes' if product['parent_product__is_lead_time_applicable'] else 'No')
        row.append(product['parent_product__discounted_life_percent'])
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    info_logger.info("Parent Data Sample Excel File has been Successfully Downloaded")
    return response


def set_child_data_sample_excel_file(request, *args):
    """
    This function will return an Sample Excel File in xlsx format which can be used for uploading the master_data
    """
    category_id = request.GET['category_id']
    category = Category.objects.values('category_name').filter(id=int(category_id))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-child_data_sample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['sku_id', 'sku_name', 'ean', 'mrp', 'weight_unit', 'weight_value',
               'status', 'repackaging_type', 'source_sku_id', 'source_sku_name', 'raw_material',
               'wastage', 'fumigation', 'label_printing', 'packing_labour', 'primary_pm_cost',
               'secondary_pm_cost', 'final_fg_cost', 'conversion_cost']
    mandatory_columns = ['sku_id', 'sku_name', 'status']
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        if column_title in mandatory_columns:
            ft = Font(color="000000FF", name="Arial", b=True)
        else:
            ft = Font(color="00000000", name="Arial", b=True)
        cell.font = ft

    products = Product.objects.values('id', 'product_sku', 'product_name', 'product_ean_code',
                                      'product_mrp', 'weight_unit', 'weight_value', 'status', 'repackaging_type') \
        .filter(
        Q(parent_product__parent_product_pro_category__category__category_name__icontains=category[0]['category_name']))
    for product in products:
        row = []
        row.append(product['product_sku'])
        row.append(product['product_name'])
        row.append(product['product_ean_code'])
        row.append(product['product_mrp'])
        row.append(product['weight_unit'])
        row.append(product['weight_value'])
        row.append(product['status'])
        row.append(product['repackaging_type'])
        source_sku_name = Repackaging.objects.select_related('source_sku').filter(destination_sku=product['id'])
        source_sku_ids = []
        source_sku_names = []
        for sourceSKU in source_sku_name:
            if sourceSKU.source_sku.product_sku not in source_sku_ids:
                source_sku_ids.append(sourceSKU.source_sku.product_sku)
            if sourceSKU.source_sku.product_name not in source_sku_names:
                source_sku_names.append(sourceSKU.source_sku.product_name)
        if source_sku_ids:
            row.append(str(source_sku_ids))
        else:
            row.append('')
        if source_sku_names:
            row.append(str(source_sku_names))
        else:
            row.append('')
        costs = DestinationRepackagingCostMapping.objects.values('raw_material', 'wastage', 'fumigation',
                                                                 'label_printing', 'packing_labour', 'primary_pm_cost',
                                                                 'secondary_pm_cost', 'final_fg_cost',
                                                                 'conversion_cost').filter(destination=product['id'])
        for cost in costs:
            row.append(cost['raw_material'])
            row.append(cost['wastage'])
            row.append(cost['fumigation'])
            row.append(cost['label_printing'])
            row.append(cost['packing_labour'])
            row.append(cost['primary_pm_cost'])
            row.append(cost['secondary_pm_cost'])
            row.append(cost['final_fg_cost'])
            row.append(cost['conversion_cost'])
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    info_logger.info("Child Data Sample Excel File has been Successfully Downloaded")
    return response


def category_sub_category_mapping_sample_excel_file(request):
    """
    This function will return an Sample Excel File in xlsx format which can be used for Mapping of
    Sub Category and Category
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-subCategory-CategorySample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['category_id', 'category_name', 'sub_category_id', 'sub_category_name', ]
    mandatory_columns = ['category_id', 'category_name']

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        if column_title in mandatory_columns:
            ft = Font(color="000000FF", name="Arial", b=True)
        else:
            ft = Font(color="00000000", name="Arial", b=True)
        cell.font = ft

    products = Category.objects.values('id', 'category_name', 'category_parent_id', 'category_parent__category_name')
    for product in products:
        row = []
        if product['category_parent_id']:
            row.append(product['category_parent_id'])
            row.append(product['category_parent__category_name'])
            row.append(product['id'])
            row.append(product['category_name'])
        else:
            row.append(product['id'])
            row.append(product['category_name'])
            row.append(product['category_parent_id'])
            row.append(product['category_parent__category_name'])
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    info_logger.info("Category and Sub Category Mapping Sample Excel File has been Successfully Downloaded")
    return response


def brand_sub_brand_mapping_sample_excel_file(request):
    """
    This function will return an Sample Excel File in xlsx format which can be used for Mapping of
    Sub Brand and Brand
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response[
        'Content-Disposition'] = 'attachment; filename={date}-subBrand-BrandMappingSample.xlsx'.format(
        date=datetime.datetime.now().strftime('%d_%b_%y_%I_%M'),
    )
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    # Sheet header, first row
    row_num = 1

    columns = ['brand_id', 'brand_name', 'sub_brand_id', 'sub_brand_name', ]
    mandatory_columns = ['brand_id', 'brand_name']

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        if column_title in mandatory_columns:
            ft = Font(color="000000FF", name="Arial", b=True)
        else:
            ft = Font(color="00000000", name="Arial", b=True)
        cell.font = ft

    products = Brand.objects.values('id', 'brand_name', 'brand_parent_id', 'brand_parent__brand_name')
    for product in products:
        row = []
        if product['brand_parent_id']:
            row.append(product['brand_parent_id'])
            row.append(product['brand_parent__brand_name'])
            row.append(product['id'])
            row.append(product['brand_name'])

        else:
            row.append(product['id'])
            row.append(product['brand_name'])
            row.append(product['brand_parent_id'])
            row.append(product['brand_parent__brand_name'])

        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    info_logger.info("Brand and Sub Brand Mapping Sample Excel File has been Successfully Downloaded")
    return response


def upload_master_data_view(request):
    """
    This function will be used for following operations:
    a)Set the Status to "Deactivated" for a Product
    b)Mapping of "Sub Brand" to "Brand"
    c)Mapping of "Sub Category" to "Category"
    d)Set the data for "Parent SKU"
    e)Mapping of Child SKU to Parent SKU
    f)Set the Child SKU Data

    After following operations, an entry will be created in 'BulkUploadForProductAttributes' Table
    """
    if request.method == 'POST':
        excel_file = request.FILES['file']
        if excel_file.name[-5:] == '.xlsx':
            data = xlsx_get(excel_file)
            form = UploadMasterDataAdminForm(request.POST, request.FILES, data)

        if not excel_file.name[-5:] == '.xlsx':
            form = UploadMasterDataAdminForm(request.POST, request.FILES)

        if form.errors:
            return render(request, 'admin/products/upload-master-data.html', {'form': form})

        if form.is_valid():
            excel_file_data = xlsx_get(excel_file)['Users']
            header_list = excel_file_data.pop(0)  # remove the header from the list
            excel_file_headers = [str(ele).lower() for ele in
                                  header_list]  # Converting headers into lowercase
            excel_file_list = []  # It will be a list of dictionaries with format [{'sku_id': 'NDPPROMAG00000018', ...}]
            excel_dict = {}
            count = 0
            for row in excel_file_data:
                for ele in row:
                    excel_dict[excel_file_headers[count]] = ele
                    count += 1
                excel_file_list.append(excel_dict)
                excel_dict = {}
                count = 0

            if request.POST['upload_master_data'] == 'master_data':
                SetMasterData.set_master_data(excel_file_list)
            if request.POST['upload_master_data'] == 'inactive_status':
                UploadMasterData.set_inactive_status(excel_file_list)
            if request.POST['upload_master_data'] == 'sub_brand_with_brand':
                UploadMasterData.set_sub_brand_and_brand(excel_file_list)
            if request.POST['upload_master_data'] == 'sub_category_with_category':
                UploadMasterData.set_sub_category_and_category(excel_file_list)
            if request.POST['upload_master_data'] == 'child_parent':
                UploadMasterData.set_child_parent(excel_file_list)
            if request.POST['upload_master_data'] == 'child_data':
                UploadMasterData.set_child_data(excel_file_list)
            if request.POST['upload_master_data'] == 'parent_data':
                UploadMasterData.set_parent_data(excel_file_list)

            attribute_id = BulkUploadForProductAttributes.objects.values('id').last()
            if attribute_id:
                request.FILES['file'].name = request.POST['upload_master_data'] + '-' + str(
                    attribute_id['id'] + 1) + '.xlsx'
            else:
                request.FILES['file'].name = request.POST['upload_master_data'] + '-' + str(1) + '.xlsx'
            product_attribute = BulkUploadForProductAttributes.objects.create(file=request.FILES['file'],
                                                                              updated_by=request.user)
            product_attribute.save()
            # return HttpResponseRedirect('/admin/products/bulkuploadforproductattributes/')
            return render(request, 'admin/products/upload-master-data.html',
                          {'form': form,
                           'success': 'Master Data Uploaded Successfully!', })

    else:
        form = UploadMasterDataAdminForm()
    return render(request, 'admin/products/upload-master-data.html', {'form': form})


def FetchDefaultChildDdetails(request):
    parent_product_id = request.GET.get('parent')
    data = {
        'found': False
    }
    if not parent_product_id:
        return JsonResponse(data)
    def_child = Product.objects.filter(parent_product=parent_product_id,
                                       reason_for_child_sku__icontains='default').last()
    if def_child:
        data = {
            'found': True,
            'product_name': def_child.product_name,
            'product_ean_code': def_child.product_ean_code,
            'product_mrp': def_child.product_mrp,
            'weight_value': def_child.weight_value,
            'weight_unit': {
                'option': def_child.weight_unit,
                'text': 'Gram'
            },
            'enable_use_parent_image_check': True if def_child.parent_product.parent_product_pro_image.exists() else False
        }

    return JsonResponse(data)


class ParentProductsAutocompleteView(AutocompleteJsonView):
    def get_queryset(self):
        queryset = ParentProduct.objects.all().order_by('name')

        if self.term:
            queryset = queryset.filter(Q(name__icontains=self.term) | Q(parent_id__icontains=self.term))

        return queryset


def FetchAllParentCategories(request):
    data = {'categories': []}
    categories = Category.objects.all()
    for category in categories:
        data['categories'].append(category.category_name)

    return JsonResponse(data, safe=False)


def FetchAllParentCategoriesWithID(request):
    data = {'categories': []}
    categories = Category.objects.all()
    for category in categories:
        data['categories'].append(category.category_name + "@" + str(category.id))

    return JsonResponse(data, safe=False)


def FetchAllProductBrands(request):
    data = {'brands': []}
    brands = Brand.objects.all()
    for brand in brands:
        data['brands'].append(brand.brand_name)

    return JsonResponse(data, safe=False)


def FetchProductDdetails(request):
    product_id = request.GET.get('product')
    data = {
        'found': False
    }
    if not product_id:
        return JsonResponse(data)
    def_product = Product.objects.filter(pk=product_id).last()
    selling_price = None
    selling_price_per_saleable_unit = None
    if def_product:
        is_ptr_applicable = def_product.parent_product.is_ptr_applicable
        case_size = def_product.parent_product.inner_case_size
        if is_ptr_applicable:
            selling_price = round(get_selling_price(def_product), 2)
            selling_price_per_saleable_unit = round(selling_price * case_size, 2)
        data = {
            'found': True,
            'product_mrp': def_product.product_mrp,
            'selling_price_per_piece': selling_price,
            'selling_price_per_saleable_unit': selling_price_per_saleable_unit
        }

    return JsonResponse(data)


def FetchDiscountedProductdetails(request):
    product_id = request.GET.get('product')
    seller_shop = request.GET.get('seller_shop')
    print(product_id, seller_shop)
    data = {
        'found': 0,
    }
    if not product_id:
        return JsonResponse(data)
    def_product = Product.objects.filter(pk=product_id).last()
    if not seller_shop:
        if def_product:
            data = {
                'found': 1,
                'product_mrp': def_product.product_mrp,
                'manual_update': def_product.is_manual_price_update
            }
        return JsonResponse(data)
    shop = Shop.objects.filter(pk=seller_shop).last()
    selling_price = None
    if def_product and shop:
        original_prod = def_product.product_ref
        bin_inventory = BinInventory.objects.filter(sku=def_product,
                                                    inventory_type__inventory_type='normal', quantity__gt=0).last()
        print(bin_inventory)
        if not bin_inventory:
            data = {
                'error': True,
                'message': "This discounted product has no inventory. Select another product"
            }
            return JsonResponse(data)
        latest_in = In.objects.filter(sku=bin_inventory.sku.product_ref, batch_id=bin_inventory.batch_id).order_by(
            '-modified_at')
        expiry_date = latest_in[0].expiry_date
        manufacturing_date = latest_in[0].manufacturing_date
        product_life = expiry_date - manufacturing_date
        remaining_life = expiry_date - datetime.date.today()
        discounted_life = math.floor(product_life.days * original_prod.parent_product.discounted_life_percent / 100)
        product_price = original_prod.product_pro_price.filter(seller_shop=shop,
                                                               approval_status=ProductPrice.APPROVED).last()
        if not product_price:
            data = {
                'error': True,
                'message': "No product price for this shop"
            }
            return JsonResponse(data)
        base_price_slab = product_price.price_slabs.filter(end_value=0).last()
        base_price = base_price_slab.ptr

        half_life = (discounted_life - 2) / 2

        if remaining_life.days <= half_life:
            selling_price = round(base_price * int(get_config('DISCOUNTED_HALF_LIFE_PRICE_PERCENT', 50)) / 100, 2)
        else:
            selling_price = round(base_price * int(get_config('DISCOUNTED_PRICE_PERCENT', 75)) / 100, 2)
        data = {
            'found': 2,
            'product_mrp': def_product.product_mrp,
            'selling_price': selling_price,
            'manual_update': def_product.is_manual_price_update
            # 'selling_price_per_saleable_unit' : selling_price_per_saleable_unit
        }
    print(data)
    return JsonResponse(data)


def get_selling_price(def_product):
    selling_price = 0
    ptr_percent = def_product.parent_product.ptr_percent
    ptr_type = def_product.parent_product.ptr_type
    if ptr_type == ParentProduct.PTR_TYPE_CHOICES.MARK_UP:
        selling_price = def_product.product_mrp / (1 + (ptr_percent / 100))
    elif ptr_type == ParentProduct.PTR_TYPE_CHOICES.MARK_DOWN:
        selling_price = def_product.product_mrp * (1 - (ptr_percent / 100))
    return selling_price


class ProductCategoryMapping(View):

    def validate_row(self, first_row, row):
        if not row[0]:
            raise Exception("{} is requied".format(first_row[0]))
        if not row[1]:
            raise Exception("{} is requied".format(first_row[1]))

    def update_mapping(self, request, file):
        try:
            with transaction.atomic():
                reader = csv.reader(codecs.iterdecode(file, 'utf-8', errors='ignore'))
                first_row = next(reader)
                for row_id, row in enumerate(reader):
                    self.validate_row(first_row, row)
                    ProductCategory.objects.filter(
                        product=Product.objects.get(product_gf_code=row[0])
                    ).update(category=Category.objects.get(id=row[1]))

                messages.success(request, 'Category Mapping updated successfully')

        except Exception as e:
            messages.error(request, "{} at Row[{}]".format(e, row_id + 2))

    @method_decorator(permission_required('products.change_product'))
    def get(self, request):
        form = ProductCategoryMappingForm()
        return render(request, 'admin/products/productcategorymapping.html',
                      {'form': form})

    @method_decorator(permission_required('products.change_product'))
    def post(self, request):
        form = ProductCategoryMappingForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data.get('file')
            self.update_mapping(request, file)

        return render(request, 'admin/products/productcategorymapping.html',
                      {'form': form})


def product_category_mapping_sample(self):
    filename = "product_category_mapping_sample.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerows([['gf_code', 'category_id'], ['GF01641', '161']])
    return response


class CityAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        buyer_shop = self.forwarded.get('buyer_shop', None)
        state = self.forwarded.get('state', None)
        qs = City.objects.all()
        if buyer_shop:
            qs = qs.filter(city_address__shop_name_id=buyer_shop,
                           city_address__address_type='shipping')
        if state:
            qs = qs.filter(state=state)
        if self.q:
            qs = qs.filter(city_name__icontains=self.q)
        return qs.distinct('city_name')


class RetailerAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Shop.objects.filter(shop_type__shop_type__in=['r', 'f'])
        if self.q:
            qs = qs.filter(shop_name__icontains=self.q)
        return qs


class SellerShopAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Shop.objects.filter(shop_type__shop_type='sp')
        if self.q:
            qs = qs.filter(shop_name__icontains=self.q)
        return qs


class ProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Product.objects.filter(product_type=Product.PRODUCT_TYPE_CHOICE.NORMAL)

        if self.q:
            qs = qs.filter(Q(product_name__icontains=self.q) |
                           Q(product_sku__icontains=self.q))
        return qs


class SourceProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Product.objects.filter(repackaging_type='source')
        if self.q:
            qs = qs.filter(Q(product_name__icontains=self.q) |
                           Q(product_sku__icontains=self.q))
        return qs


class PincodeAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        city = self.forwarded.get('city', None)
        buyer_shop = self.forwarded.get('buyer_shop', None)
        qs = Pincode.objects.all()
        if buyer_shop:
            qs = qs.filter(pincode_address__shop_name_id=buyer_shop,
                           pincode_address__address_type='shipping').distinct('pincode')
            return qs
        if city:
            qs = qs.filter(city_id=city)
        return qs.distinct('pincode')


class ProductPriceUpload(View):
    form_class = NewProductPriceUpload
    template_name = 'admin/products/NewProductPriceUpload.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def products_price_qs(self, data):
        qs = ProductPrice.objects.filter(seller_shop=data['seller_shop'])
        if data['city']:
            qs = qs.filter(city=data['city'])
        if data['pincode_from'] and data['pincode_to']:
            pincode_range = [i for i in range(int(data['pincode_from']),
                                              int(data['pincode_to']))]
            qs = qs.filter(pincode__in=pincode_range)
        if data['buyer_shop']:
            qs = qs.filter(buyer_shop=data['buyer_shop'])
        if data['product']:
            qs = qs.filter(product=data['product'])
        return qs.values_list(
            'product__product_sku', 'product__product_name',
            'seller_shop__shop_name', 'product__product_mrp',
            'selling_price', 'city_id', 'city__city_name', 'pincode__pincode',
            'buyer_shop_id', 'buyer_shop__shop_name', 'start_date', 'end_date',
            'approval_status')

    def validate_row(self, first_row, row, mrp_col_present=True):
        # if (row[0] and not re.match("^[\d]*$", str(row[0]))) or not row[0]:
        #     raise Exception("{} - Please enter a valid {}"
        #                     "".format(row[0], first_row[0]))
        if mrp_col_present:
            # if ((row[3] and not re.match("^\d{0,8}(\.\d{1,2})?$", str(row[3]))) or
            #         not row[3]):
            #     raise Exception("{} - Please enter a valid {}"
            #                     "".format(row[3], first_row[3]))
            if ((row[4] and not re.match("^\d{0,8}(\.\d{1,2})?$", str(row[4]))) or
                    not row[4]):
                raise Exception("{} - Please enter a valid {}"
                                "".format(row[4], first_row[4]))
            if not row[10]:
                raise Exception("{} - Please enter a valid {}"
                                "".format(row[10], first_row[10]))
            # if not row[11]:
            #     raise Exception("{} - Please enter a valid {}"
            #                     "".format(row[11], first_row[11]))
        else:
            if ((row[3] and not re.match("^\d{0,8}(\.\d{1,2})?$", str(row[3]))) or
                    not row[3]):
                raise Exception("{} - Please enter a valid {}"
                                "".format(row[3], first_row[3]))
            # if ((row[4] and not re.match("^\d{0,8}(\.\d{1,2})?$", str(row[4]))) or
            #         not row[4]):
            #     raise Exception("{} - Please enter a valid {}"
            #                     "".format(row[4], first_row[4]))
            if not row[9]:
                raise Exception("{} - Please enter a valid {}"
                                "".format(row[9], first_row[9]))
            if not row[10]:
                raise Exception("{} - Please enter a valid {}"
                                "".format(row[10], first_row[10]))

    def create_product_price(self, request, data):
        try:
            with transaction.atomic():
                wb_obj = openpyxl.load_workbook(data.get('csv_file'))
                sheet_obj = wb_obj.active
                first_row = next(sheet_obj.iter_rows(values_only=True))
                mrp_col_present = False
                for col in first_row:
                    if 'mrp' in col.lower():
                        mrp_col_present = True
                for row_id, row in enumerate(sheet_obj.iter_rows(
                        min_row=2, max_row=None, min_col=None, max_col=None,
                        values_only=True
                )):
                    self.validate_row(first_row, row, mrp_col_present)
                    product = Product.objects.get(product_sku=row[0])
                    if not product.product_mrp:
                        raise Exception("Product MRP not present at Child Product level")
                    if mrp_col_present:
                        if row[7] and Pincode.objects.filter(Q(pincode=row[7]) | Q(id=row[7])).exists():
                            # pincode = Pincode.objects.values('id').get(pincode=row[7])['id']
                            pincode = Pincode.objects.filter(Q(pincode=row[7]) | Q(id=row[7])).last()
                        else:
                            pincode = None
                        ProductPrice.objects.create(
                            product=product, mrp=product.product_mrp,
                            selling_price=round(Decimal(row[4]), 2),
                            seller_shop_id=int(data['seller_shop'].id),
                            buyer_shop_id=int(row[8]) if row[8] else None,
                            city_id=int(row[5]) if row[5] else None,
                            pincode=pincode,
                            start_date=row[10],
                            approval_status=ProductPrice.APPROVED)
                    else:
                        if row[6] and Pincode.objects.filter(Q(pincode=row[6]) | Q(id=row[6])).exists():
                            # pincode = Pincode.objects.values('id').get(pincode=row[6])['id']
                            pincode = Pincode.objects.filter(Q(pincode=row[6]) | Q(id=row[6])).last()
                        else:
                            pincode = None
                        ProductPrice.objects.create(
                            product=product, mrp=product.product_mrp,
                            selling_price=round(Decimal(row[3]), 2),
                            seller_shop_id=int(data['seller_shop'].id),
                            buyer_shop_id=int(row[7]) if row[7] else None,
                            city_id=int(row[4]) if row[4] else None,
                            pincode=pincode,
                            start_date=row[9],
                            approval_status=ProductPrice.APPROVED)

                messages.success(request, 'Prices uploaded successfully')

        except Exception as e:
            messages.error(request, "{} at Row[{}] for {}"
                                    "".format(e, row_id + 2, row[1]))

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            # if action is upload
            if data['action'] == '1':
                self.create_product_price(request, data)
            # if action is download
            elif data['action'] == '2':
                return products_price_excel(self.products_price_qs(data))

        return render(request, self.template_name, {'form': form})


class VendorAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self, *args, **kwargs):
        qs = Vendor.objects.none
        if self.q:
            qs = Vendor.objects.filter(
                Q(vendor_name__icontains=self.q)
            )
        return qs


class ProductShopAutocomplete(autocomplete.Select2QuerySetView):

    def get_queryset(self, *args, **kwargs):
        seller_shop = self.forwarded.get('seller_shop', None)
        qs = []
        if seller_shop:
            ids = Product.objects.filter(repackaging_type='source', product_pro_price__seller_shop=seller_shop) \
                .values_list('id', flat=True).distinct()
            normal_type = InventoryType.objects.filter(inventory_type='normal').last()
            product_list = get_stock(seller_shop, normal_type, ids)
            product_list = {k: v for k, v in product_list.items() if v > 0}
            qs = Product.objects.filter(id__in=product_list.keys())
            if self.q:
                qs = qs.filter(product_name__icontains=self.q)
        return qs


class SourceRepackageDetail(View):

    def get(self, *args, **kwargs):
        product_id = self.request.GET.get('sku_id')
        shop_id = self.request.GET.get('shop_id')
        product_obj = Product.objects.values('weight_value', 'product_sku').get(id=product_id)

        if product_obj['weight_value'] is None:
            return JsonResponse({"success": False, "error": "Source SKU Weight Value Not Found"})

        try:
            normal_type = InventoryType.objects.filter(inventory_type='normal').last()
            product_inv = get_stock(shop_id, normal_type, [int(product_id)])
            if product_inv and int(product_id) in product_inv:
                source_quantity = product_inv[int(product_id)]
                if source_quantity <= 0:
                    return JsonResponse({"success": False, "error": "Source Not Available In Warehouse"})
            else:
                return JsonResponse({"success": False, "error": "Warehouse Inventory Does Not Exist"})
        except Exception as e:
            return JsonResponse({"success": False, "error": "Warehouse Inventory Could not be fetched"})

        return JsonResponse({
            "available_weight": (source_quantity * product_obj['weight_value']) / 1000,
            "source_sku_weight": product_obj['weight_value'] / 1000,
            "available_source_quantity": source_quantity,
            "success": True})


class DestinationProductAutocomplete(autocomplete.Select2QuerySetView):

    def get_queryset(self, *args, **kwargs):
        source_sku = self.forwarded.get('source_sku', None)
        qs = []
        if source_sku:
            psm = ProductSourceMapping.objects.filter(source_sku=source_sku, status=True).values('destination_sku')
            qs = Product.objects.filter(id__in=psm)
            if self.q:
                qs = qs.filter(product_name__icontains=self.q)
        return qs


def products_export_for_vendor(request, id=None):
    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + "product_list.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    vendor_id = request.GET.get('id', None)
    vendor_mapped_product = ProductVendorMapping.objects.filter(vendor=vendor_id)

    writer.writerow(
        ['id', 'product_name', 'product_sku', 'mrp', 'brand_to_gram_price_unit', 'brand_to_gram_price', 'case_size'])
    if vendor_mapped_product:
        product_id = ProductVendorMapping.objects.filter(vendor=vendor_id).values('product')
        products = Product.objects.filter(status="active").exclude(id__in=product_id).only('id', 'product_name',
                                                                                           'product_sku', 'product_mrp')
        for product in products:
            writer.writerow(
                [product.id, product.product_name, product.product_sku, '', '', '', product.product_case_size])
    else:
        products = Product.objects.filter(status="active").only('id', 'product_name', 'product_sku', 'product_mrp')
        for product in products:
            writer.writerow(
                [product.id, product.product_name, product.product_sku, '', '', '', product.product_case_size])

    return response


def all_product_mapped_to_vendor(request, id=None):
    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + "product_vendor_list.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    vendor_id = request.GET.get('id', None)
    vendor_mapped_product = ProductVendorMapping.objects.filter(vendor=vendor_id)
    writer.writerow(
        ['id', 'product_name', 'product_sku', 'mrp', 'brand_to_gram_price_unit', 'brand_to_gram_price', 'case_size'])
    if vendor_mapped_product:
        products_vendors = ProductVendorMapping.objects.filter(vendor=vendor_id).only('product', 'vendor',
                                                                                      'brand_to_gram_price_unit',
                                                                                      'product_price',
                                                                                      'product_price_pack', 'case_size')
        for product_vendor in products_vendors:
            if product_vendor.status == True:
                if product_vendor.brand_to_gram_price_unit == "Per Piece":
                    writer.writerow([product_vendor.product.id, product_vendor.product.product_name,
                                     product_vendor.product.product_sku, product_vendor.product_mrp,
                                     product_vendor.brand_to_gram_price_unit, product_vendor.product_price,
                                     product_vendor.case_size])
                else:
                    writer.writerow([product_vendor.product.id, product_vendor.product.product_name,
                                     product_vendor.product.product_sku, product_vendor.product_mrp,
                                     product_vendor.brand_to_gram_price_unit, product_vendor.product_price_pack,
                                     product_vendor.case_size])
    return response


def bulk_product_vendor_csv_upload_view(request):
    all_vendors = Vendor.objects.all()

    if request.method == 'POST':
        form = BulkProductVendorMapping(request.POST, request.FILES)

        if form.errors:
            return render(request, 'admin/products/bulk-upload-vendor-details.html', {
                'form': form,
                'all_vendor': all_vendors.values(),
            })
        if form.is_valid():
            upload_file = form.cleaned_data.get('file')
            vendor_id = request.POST.get('select')
            reader = csv.reader(codecs.iterdecode(upload_file, 'utf-8', errors='ignore'))
            first_row = next(reader)
            try:
                for row_id, row in enumerate(reader):
                    if len(row) == 0:
                        continue
                    if '' in row:
                        if (row[0] == '' and row[1] == '' and row[2] == '' and row[3] == '' and row[4] == '' and
                                row[5] == '' and row[6] == ''):
                            continue

                    if row[4].title() == "Per Piece":
                        product_vendor = ProductVendorMapping.objects.create(
                            vendor=Vendor.objects.get(id=vendor_id),
                            product=Product.objects.get(id=row[0]),
                            product_mrp=row[3],
                            brand_to_gram_price_unit=row[4].title(),
                            product_price=row[5],
                            case_size=row[6],
                        )
                    else:
                        product_vendor = ProductVendorMapping.objects.create(
                            vendor=Vendor.objects.get(id=vendor_id),
                            product=Product.objects.get(id=row[0]),
                            product_mrp=row[3],
                            brand_to_gram_price_unit=row[4].title(),
                            product_price_pack=row[5],
                            case_size=row[6],
                        )
                    product_vendor.save()
            except Exception as e:
                print(e)
            return render(request, 'admin/products/bulk-upload-vendor-details.html', {
                'form': form,
                'all_vendor': all_vendors.values(),
                'success': 'Product Vendor Mapping CSV uploaded successfully !',
            })
    else:
        form = BulkProductVendorMapping()
    return render(request, 'admin/products/bulk-upload-vendor-details.html',
                  {'all_vendor': all_vendors.values(), 'form': form})


def get_slab_product_price_sample_csv(request):
    """
    returns sample CSV for bulk creation of Slab Product Prices
    """
    filename = "slab_product_price_sample_csv.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(["SKU", "Product Name", "Shop Id", "Shop Name", "MRP", "Slab 1 Qty", "Selling Price 1",
                     "Offer Price 1", "Offer Price 1 Start Date(dd-mm-yy)", "Offer Price 1 End Date(dd-mm-yy)",
                     "Slab 2 Qty", "Selling Price 2", "Offer Price 2", "Offer Price 2 Start Date(dd-mm-yy)",
                     "Offer Price 2 End Date(dd-mm-yy)"])
    writer.writerow(["BDCHNKDOV00000001", "Dove CREAM BAR 100G shop", "600",
                     "GFDN SERVICES PVT LTD (NOIDA) - 9319404555 - Rakesh Kumar - Service Partner", "47", "9", "46",
                     "45.5", "01-03-21", "30-04-21", "10", "45", "44.5", "01-03-21", "30-04-21"])
    return response


def get_discounted_product_price_sample_csv(request):
    """
    returns sample CSV for bulk creation of Slab Product Prices
    """
    filename = "discounted_product_price_sample_csv.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(["SKU", "Product Name", "Shop Id", "Shop Name", "selling price"])
    writer.writerow(
        ["DRGRSNGDAW00000020", "Daawat Rozana Super, 5 KG", "600", "GFDN SERVICES PVT LTD (DELHI)", "123.00"])
    return response


def slab_product_price_csv_upload(request):
    """
    Creates Slab Product Prices in bulk through CSV upload
    """
    if request.method == 'POST':
        form = UploadSlabProductPriceForm(request.POST, request.FILES)

        if form.errors:
            return render(request, 'admin/products/bulk-slab-product-price.html', {'form': form})

        if form.is_valid():
            upload_file = form.cleaned_data.get('file')
            reader = csv.reader(codecs.iterdecode(upload_file, 'utf-8', errors='ignore'))
            first_row = next(reader)

            try:
                for row_id, row in enumerate(reader):
                    product = Product.objects.filter(product_sku=row[0]).last()
                    seller_shop_id = int(row[2])

                    # Create ProductPrice
                    product_price = SlabProductPrice(product=product, mrp=product.product_mrp,
                                                     seller_shop_id=seller_shop_id)
                    product_price.save()

                    # Get selling price applicable for first price slab
                    is_ptr_applicable = product.parent_product.is_ptr_applicable
                    if is_ptr_applicable:
                        selling_price = get_selling_price(product)
                    else:
                        selling_price = float(row[6])

                    # Create Price Slabs

                    # Create Price Slab 1
                    price_slab_1 = PriceSlab(product_price=product_price, start_value=0, end_value=int(row[5]),
                                             selling_price=selling_price)
                    if row[7]:
                        price_slab_1.offer_price = float(row[7])
                        price_slab_1.offer_price_start_date = getStrToDate(row[8], '%d-%m-%y').strftime('%Y-%m-%d')
                        price_slab_1.offer_price_end_date = getStrToDate(row[9], '%d-%m-%y').strftime('%Y-%m-%d')
                    price_slab_1.save()

                    # If slab 1 quantity is Zero then slab is not to be created
                    if int(row[5]) == 0:
                        continue
                    # Create Price Slab 2
                    price_slab_2 = PriceSlab(product_price=product_price, start_value=row[10], end_value=0,
                                             selling_price=float(row[11]))
                    if row[12]:
                        price_slab_2.offer_price = float(row[12])
                        price_slab_2.offer_price_start_date = getStrToDate(row[13], '%d-%m-%y').strftime('%Y-%m-%d')
                        price_slab_2.offer_price_end_date = getStrToDate(row[14], '%d-%m-%y').strftime('%Y-%m-%d')
                    price_slab_2.save()

            except Exception as e:
                print(e)
                msg = 'Unable to create price for row {}'.format(row_id + 1)
                return render(request, 'admin/products/bulk-slab-product-price.html', {'form': form, 'error': msg})

            return render(request, 'admin/products/bulk-slab-product-price.html', {
                'form': form,
                'success': 'Slab Product Prices uploaded successfully !',
            })
    else:
        form = UploadSlabProductPriceForm()
    return render(request, 'admin/products/bulk-slab-product-price.html', {'form': form})


def discounted_product_price_csv_upload(request):
    """
    Creates Discounted Product Prices in bulk through CSV upload
    """
    if request.method == 'POST':
        form = UploadDiscountedProductPriceForm(request.POST, request.FILES)

        if form.errors:
            return render(request, 'admin/products/bulk-discounted-product-price.html', {'form': form})

        if form.is_valid():
            upload_file = form.cleaned_data.get('file')
            reader = csv.reader(codecs.iterdecode(upload_file, 'utf-8', errors='ignore'))
            first_row = next(reader)

            try:
                for row_id, row in enumerate(reader):
                    with transaction.atomic():
                        product = Product.objects.filter(product_sku=row[0]).last()
                        seller_shop_id = int(row[2])
                        seller_shop = Shop.objects.filter(pk=seller_shop_id).last()

                        if not product.is_manual_price_update:
                            original_prod = product.product_ref
                            bin_inventory = BinInventory.objects.filter(sku=product,
                                                                        inventory_type__inventory_type='normal',
                                                                        quantity__gt=0).last()
                            latest_in = In.objects.filter(sku=bin_inventory.sku.product_ref,
                                                          batch_id=bin_inventory.batch_id).order_by('-modified_at')
                            expiry_date = latest_in[0].expiry_date
                            manufacturing_date = latest_in[0].manufacturing_date
                            product_life = expiry_date - manufacturing_date
                            remaining_life = expiry_date - datetime.date.today()
                            discounted_life = math.floor(
                                product_life.days * original_prod.parent_product.discounted_life_percent / 100)
                            product_price = original_prod.product_pro_price.filter(seller_shop=seller_shop,
                                                                                   approval_status=ProductPrice.APPROVED).last()
                            base_price_slab = product_price.price_slabs.filter(end_value=0).last()
                            base_price = base_price_slab.ptr

                            half_life = (discounted_life - 2) / 2

                            if remaining_life.days <= half_life:
                                selling_price = round(
                                    base_price * int(get_config('DISCOUNTED_HALF_LIFE_PRICE_PERCENT', 50)) / 100, 2)
                            else:
                                selling_price = round(
                                    base_price * int(get_config('DISCOUNTED_PRICE_PERCENT', 75)) / 100, 2)

                        else:
                            selling_price = float(row[4])

                        # Create Product Price
                        product_price = DiscountedProductPrice(product=product, mrp=product.product_mrp,
                                                               seller_shop_id=seller_shop_id,
                                                               selling_price=selling_price,
                                                               start_date=datetime.datetime.today(),
                                                               approval_status=ProductPrice.APPROVED)
                        product_price.save()

                        # Create Price for discounted Product
                        discounted_product_price = PriceSlab(product_price=product_price, selling_price=selling_price,
                                                             start_value=1, end_value=0)
                        discounted_product_price.save()

            except Exception as e:
                print(e)
                msg = 'Unable to create price for row {}'.format(row_id + 1)
                return render(request, 'admin/products/bulk-discounted-product-price.html',
                              {'form': form, 'error': msg})

            return render(request, 'admin/products/bulk-discounted-product-price.html', {
                'form': form,
                'success': 'Slab Product Prices uploaded successfully !',
            })
    else:
        form = UploadDiscountedProductPriceForm()
    return render(request, 'admin/products/bulk-discounted-product-price.html', {'form': form})


class PackingProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Product.objects.filter(repackaging_type='packing_material')
        if self.q:
            qs = qs.filter(Q(product_name__icontains=self.q) |
                           Q(product_sku__icontains=self.q))
        return qs


class PackingMaterialCheck(View):

    def get(self, *args, **kwargs):
        try:
            ProductPackingMapping.objects.get(sku_id=self.request.GET.get('sku_id'))
            return JsonResponse({"success": True})
        except:
            return JsonResponse({"success": False, "error": "Please Map A Packing Material To The Selected Destination"
                                                            " Product First"})


def packing_material_inventory(request):
    if request.method == 'POST':
        form = UploadPackingSkuInventoryAdminForm(request.POST, request.FILES)
        msg = 'Updated Successfully!'
        error = 'Something Went Wrong!'
        if form.errors:
            return render(request, 'admin/products/packing-sku-inventory.html', {'form': form})

        if form.is_valid():
            upload_data = form.cleaned_data.get('file')
            threshold = form.cleaned_data.get('inventory_threshold')
            GlobalConfig.objects.filter(key='packing_sku_inventory_threshold_kg').update(value=threshold)
            if not upload_data:
                return render(request, 'admin/products/packing-sku-inventory.html', {'form': form, 'success': msg})
            try:
                with transaction.atomic():
                    stock_movement_obj = StockMovementCSV.create_stock_movement_csv(request.user, request.FILES['file'],
                                                                                    5)
                    for row in upload_data:
                        in_quantity_dict = {}
                        out_quantity_dict = {}
                        type_normal = InventoryType.objects.filter(inventory_type='normal').last()
                        type_damaged = InventoryType.objects.filter(inventory_type='damaged').last()
                        type_expired = InventoryType.objects.filter(inventory_type='expired').last()
                        type_missing = InventoryType.objects.filter(inventory_type='missing').last()
                        # get the type of stock
                        stock_correction_type = 'stock_adjustment'
                        warehouse_id = row[0]
                        warehouse_obj = Shop.objects.get(id=warehouse_id)
                        expiry_date = row[3]
                        bin_id = row[4]
                        bin_obj = Bin.objects.filter(bin_id=bin_id, warehouse=warehouse_id).last()
                        sku = row[2]
                        product_obj = Product.objects.filter(product_sku=sku).last()
                        # create batch id
                        batch_id = create_batch_id(sku, expiry_date)
                        normal_wt = float(row[5]) * 1000
                        damaged_wt = float(row[6]) * 1000
                        expired_wt = float(row[7]) * 1000
                        missing_wt = float(row[8]) * 1000

                        in_quantity_dict, out_quantity_dict = update_inv_dict(warehouse_obj, bin_obj, product_obj,
                                                                              batch_id, type_normal, normal_wt,
                                                                              in_quantity_dict, out_quantity_dict)

                        in_quantity_dict, out_quantity_dict = update_inv_dict(warehouse_obj, bin_obj, product_obj,
                                                                              batch_id, type_damaged, damaged_wt,
                                                                              in_quantity_dict, out_quantity_dict)

                        in_quantity_dict, out_quantity_dict = update_inv_dict(warehouse_obj, bin_obj, product_obj,
                                                                              batch_id, type_expired, expired_wt,
                                                                              in_quantity_dict, out_quantity_dict)

                        in_quantity_dict, out_quantity_dict = update_inv_dict(warehouse_obj, bin_obj, product_obj,
                                                                              batch_id, type_missing, missing_wt,
                                                                              in_quantity_dict, out_quantity_dict)

                        inventory_state = InventoryState.objects.filter(inventory_state='total_available').last()
                        transaction_type = 'stock_correction_in_type'
                        for inv_type, weight in in_quantity_dict.items():
                            manufacturing_date = get_manufacturing_date(batch_id)
                            in_obj = InCommonFunctions.create_only_in(warehouse_obj, stock_correction_type,
                                                                      stock_movement_obj[0].id, product_obj, batch_id,
                                                                      0, inv_type, weight, manufacturing_date)
                            inventory_in_and_out_weight(warehouse_obj, bin_obj, product_obj, batch_id, inv_type,
                                                        inventory_state, weight, transaction_type, in_obj.id)
                            # Create data in Stock Correction change Model
                            InternalStockCorrectionChange.create_stock_inventory_change(warehouse_obj, product_obj,
                                                                                        batch_id, bin_obj, 'In', 0,
                                                                                        stock_movement_obj[0], inv_type,
                                                                                        weight)

                        transaction_type = 'stock_correction_out_type'
                        for inv_type, weight in out_quantity_dict.items():
                            out_obj = Out.objects.create(warehouse=warehouse_obj, out_type='stock_correction_out_type',
                                                         out_type_id=stock_movement_obj[0].id, sku=product_obj,
                                                         batch_id=batch_id, quantity=0, inventory_type=inv_type,
                                                         weight=weight)
                            inventory_in_and_out_weight(warehouse_obj, bin_obj, product_obj, batch_id, inv_type,
                                                        inventory_state, weight, transaction_type, out_obj.id)
                            # Create data in Stock Correction change Model
                            InternalStockCorrectionChange.create_stock_inventory_change(warehouse_obj, product_obj,
                                                                                        batch_id, bin_obj, 'Out', 0,
                                                                                        stock_movement_obj[0], inv_type,
                                                                                        weight)
                error = ''
            except Exception as e:
                print(e)
                msg = ''
            return render(request, 'admin/products/packing-sku-inventory.html', {'form': form, 'success': msg,
                                                                                 'error': error})
    else:
        form = UploadPackingSkuInventoryAdminForm()
    return render(request, 'admin/products/packing-sku-inventory.html', {'form': form})


def update_inv_dict(warehouse, bin_obj, product_obj, batch_id, inv_type, input_wt, in_quantity_dict,
                    out_quantity_dict):
    bin_inv = BinInventory.objects.filter(warehouse=warehouse.id, bin=bin_obj, sku=product_obj, batch_id=batch_id,
                                          inventory_type__id=inv_type.id).last()
    if bin_inv:
        if bin_inv.weight <= input_wt:
            in_quantity_dict[inv_type] = input_wt
        else:
            out_quantity_dict[inv_type] = input_wt
    else:
        BinInventory.objects.get_or_create(warehouse=warehouse, bin=bin_obj,
                                           batch_id=batch_id, sku=product_obj, in_stock=True, quantity=0,
                                           weight=input_wt, inventory_type=inv_type)
        in_quantity_dict[inv_type] = input_wt

    return in_quantity_dict, out_quantity_dict


def packing_material_inventory_download(request):
    dt = datetime.datetime.now().strftime("%d_%b_%y_%H_%M_")
    filename = str(dt) + "packing_material_inventory.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)

    bin_inventories = BinInventory.objects.filter(sku__repackaging_type='packing_material').order_by('warehouse_id',
                                                                                                     'sku_id',
                                                                                                     'bin_id',
                                                                                                     'batch_id')

    headings = [
        'Warehouse ID', 'Parent ID', 'Parent Name', 'SKU ID', 'SKU Name', 'Product Status', 'Batch ID', 'Bin Id', 'MRP',
        'Normal Weight (Kg)', 'Damaged Weight (Kg)', 'Expired Weight (Kg)', 'Missing Weight (Kg)'
    ]

    writer.writerow(headings)

    row = []
    current_sku_batch = ''
    for inv in bin_inventories:
        if current_sku_batch != str(inv.warehouse_id) + str(inv.sku.id) + str(inv.bin_id) + str(inv.batch_id):
            if row:
                writer.writerow(row)
            sku = inv.sku
            current_sku_batch = str(inv.warehouse_id) + str(inv.sku.id) + str(inv.bin_id) + str(inv.batch_id)
            parent = inv.sku.parent_product
            row = [inv.warehouse_id, parent.parent_id, parent.name, sku.product_sku, sku.product_name, sku.status,
                   inv.batch_id,
                   inv.bin.bin_id, sku.product_mrp, 0, 0, 0, 0]

        if inv.inventory_type.inventory_type == 'normal':
            row[9] = inv.weight / 1000
        elif inv.inventory_type.inventory_type == 'damaged':
            row[10] = inv.weight / 1000
        elif inv.inventory_type.inventory_type == 'expired':
            row[11] = inv.weight / 1000
        elif inv.inventory_type.inventory_type == 'missing':
            row[12] = inv.weight / 1000

    writer.writerow(row)
    return response


def packing_material_inventory_sample_upload(request):
    filename = "packing_material_inventory_sample.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(['Warehouse ID', 'Product Name', 'SKU', 'Expiry Date', 'Bin Id', 'Normal Weight (Kg)',
                     'Damaged Weight (Kg)', 'Expired Weight (Kg)', 'Missing Weight (Kg)'])
    writer.writerow(['88', 'Plastic Wrap, 1000gm', 'HOKBACFRT00000021', '20/08/2022', 'V2VZ01SR001-0001',
                     '0', '0', '0', '0'])
    return response


class HSNAutocomplete(autocomplete.Select2QuerySetView):
    """

    HSN auto complete result set

    """

    def get_queryset(self, *args, **kwargs):
        qs = hsn_queryset(self)
        return qs


class DiscountedProductAutocomplete(autocomplete.Select2QuerySetView):
    """
    Get discounted product
    """

    def get_queryset(self):
        qs = Product.objects.filter(product_type=1)

        if self.q:
            qs = qs.filter(product_name__icontains=self.q)

        return qs


def franchise_po_fail_status(request, pk):
    order = get_object_or_404(Order, pk=pk)

    dt = datetime.datetime.now().strftime("%d_%b_%y_%I_%M")
    filename = str(dt) + " - OrderProductsUnavailableForPO.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    writer = csv.writer(response)
    writer.writerow(["SKU", "Name", "EAN"])
    products = order.ordered_cart.rt_cart_list.all()
    for mapp in products:
        p = mapp.cart_product
        if not RetailerProduct.objects.filter(linked_product=p, shop=order.buyer_shop).exists():
            writer.writerow([p.product_sku, p.product_name, p.product_ean_code])
    return response
