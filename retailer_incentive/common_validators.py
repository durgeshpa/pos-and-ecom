import datetime
import re
import openpyxl
from shops.models import Shop


# validation of uploaded incentive file
def bulk_incentive_data_validation(file):
    error_file_list = []
    validated_rows = []

    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    first_row = next(sheet_obj.iter_rows(values_only=True))
    error_file_list.append(list(first_row) + ["Status"])

    shops = Shop.objects.filter(shop_type__shop_type__in=['r', 'f'])
    for row_id, row in enumerate(sheet_obj.iter_rows(
            min_row=2, max_row=None, min_col=None, max_col=None,
            values_only=True
    )):

        error_msg = []
        row = list(row)
        if row == [None, None, None, None, None, None, None]:
            break
        if not row[0]:
            error_msg.append(f"{first_row[0]} cant be blank")
        elif not shops.filter(id=int(row[0])).exists():
            error_msg.append(f"{first_row[0]} is incorrect")
        else:
            row[0] = int(row[0])

        if not row[1]:
            error_msg.append(f"{first_row[1]} cant be blank")
        if row[1] and not shops.filter(shop_name__iexact=str(row[1])).exists():
            error_msg.append(f"{first_row[1]} is incorrect")

        if not row[2]:
            error_msg.append(f"{first_row[2]} cant be blank")
        elif str(row[2]).lower() not in ['yes', 'no']:
            error_msg.append(f"{first_row[2]} can only be 'Yes' or 'No' ")
        else:
            row[2] = True if str(row[2]).lower() == "yes" else False

        if not row[3]:
            error_msg.append(f"{first_row[4]} cant be blank")
        elif not re.match("^\d+[.]?[\d]{0,2}$", str(row[3])):
            error_msg.append(f"{first_row[3]} {row[3]} can only be a numeric value")
        else:
            row[3] = str(round(row[3], 2))

        if not row[4]:
            error_msg.append(f"{first_row[4]} cant be blank")
        elif isinstance(row[4], str):
            try:
                uploaded_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                row[4] = uploaded_date
            except:
                error_msg.append(f"{first_row[4]} wrong date formate {row[4]}")

        elif isinstance(row[4], datetime.datetime):
            if not row[4].date():
                error_msg.append(f"{first_row[4]} wrong date formate {row[4]}")
            else:
                row[4] = row[4].date()
        elif isinstance(row[4], datetime.date):
            if not row[4].date():
                error_msg.append(f"{first_row[4]} wrong date formate {row[4]}")
            else:
                row[4] = row[4].date()

        if not row[5]:
            error_msg.append(f"{first_row[5]} cant be blank")
        elif not re.match("^\d+[.]?[\d]{0,2}$", str(row[5])):
            error_msg.append(f"{first_row[5]} {row[5]} can only be a numeric value")
        else:
            row[5] = str(round(row[5], 2))

        if not row[6]:
            error_msg.append(f"{first_row[6]} cant be blank")
        if not re.match("^\d+[.]?[\d]{0,2}$", str(row[6])):
            error_msg.append(f"{first_row[6]} {row[6]} can only be a numeric value")
        else:
            row[6] = str(round(row[6], 2))

        if error_msg:
            msg = ", "
            msg = msg.join(map(str, error_msg))
            error_file_list.append(list(row) + [msg])
        else:
            validated_rows.append(row)

    return error_file_list, validated_rows
